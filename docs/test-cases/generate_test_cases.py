#!/usr/bin/env python3
"""Generate Treasury System BRD v3 Test Cases Excel file."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Colors ──
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
ROW_LIGHT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ROW_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Loại colors
HAPPY_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
HAPPY_FONT = Font(name="Arial", size=10, color="006100")
EDGE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
EDGE_FONT = Font(name="Arial", size=10, color="9C6500")
NEG_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NEG_FONT = Font(name="Arial", size=10, color="9C0006")

# Mức độ colors
CRIT_FILL = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
CRIT_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HIGH_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
HIGH_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
MED_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
MED_FONT = Font(name="Arial", size=10, color="333333")
LOW_FILL = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
LOW_FONT = Font(name="Arial", size=10, color="333333")

THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

COLUMNS = ["STT", "Mã TC", "Nhóm kiểm thử", "Mô tả test case", "Điều kiện đầu vào",
           "Bước thực hiện", "Kết quả mong đợi", "Loại", "Mức độ", "Ghi chú"]

COL_WIDTHS = [6, 14, 18, 35, 30, 35, 35, 12, 12, 20]


def style_sheet(ws, data_rows):
    """Apply formatting to a test case sheet."""
    # Header
    ws.freeze_panes = "A2"
    for col_idx, (col_name, width) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, row_data in enumerate(data_rows, 2):
        alt_fill = ROW_LIGHT if row_idx % 2 == 0 else ROW_WHITE
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN
            cell.font = Font(name="Arial", size=10)

            # STT column center
            if col_idx == 1:
                cell.alignment = CENTER_ALIGN

            # Loại column (col 8)
            if col_idx == 8:
                cell.alignment = CENTER_ALIGN
                if val == "Happy":
                    cell.fill = HAPPY_FILL
                    cell.font = HAPPY_FONT
                elif val == "Edge":
                    cell.fill = EDGE_FILL
                    cell.font = EDGE_FONT
                elif val == "Negative":
                    cell.fill = NEG_FILL
                    cell.font = NEG_FONT
                continue

            # Mức độ column (col 9)
            if col_idx == 9:
                cell.alignment = CENTER_ALIGN
                if val == "Critical":
                    cell.fill = CRIT_FILL
                    cell.font = CRIT_FONT
                elif val == "High":
                    cell.fill = HIGH_FILL
                    cell.font = HIGH_FONT
                elif val == "Medium":
                    cell.fill = MED_FILL
                    cell.font = MED_FONT
                elif val == "Low":
                    cell.fill = LOW_FILL
                    cell.font = LOW_FONT
                continue

            # Default alternating
            if col_idx not in (8, 9):
                cell.fill = alt_fill


def create_guide_sheet(wb):
    """Create Hướng dẫn sheet."""
    ws = wb.active
    ws.title = "Hướng dẫn"

    title_font = Font(name="Arial", size=16, bold=True, color="1F4E79")
    sub_font = Font(name="Arial", size=12, bold=True, color="1F4E79")
    body_font = Font(name="Arial", size=11)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 20

    r = 2
    ws.merge_cells(f"B{r}:D{r}")
    c = ws.cell(row=r, column=2, value="TEST CASES — HỆ THỐNG TREASURY")
    c.font = title_font

    r += 1
    ws.merge_cells(f"B{r}:D{r}")
    c = ws.cell(row=r, column=2, value="KIENLONGBANK — BRD Phiên bản 3.0")
    c.font = Font(name="Arial", size=13, color="1F4E79")

    r += 2
    info = [
        ("Mã tài liệu:", "TC-Treasury-BRD-v3"),
        ("Phiên bản BRD:", "3.0 (02/04/2026)"),
        ("Ngày tạo test case:", datetime.now().strftime("%d/%m/%Y")),
        ("Người tạo:", "KAI (AI Banking Assistant)"),
        ("Người review:", "Trần Thị Linh Phương (K.NV&ĐCTC)"),
        ("Tổng số test case:", "340+"),
        ("Tổng số sheet:", "12 (bao gồm sheet hướng dẫn)"),
    ]
    for label, val in info:
        ws.cell(row=r, column=2, value=label).font = Font(name="Arial", size=11, bold=True)
        ws.cell(row=r, column=3, value=val).font = body_font
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="CẤU TRÚC FILE").font = sub_font
    r += 1
    sheets = [
        ("M1-FX Spot/Forward", "Test case giao dịch FX Giao ngay / Kỳ hạn"),
        ("M1-FX Swap", "Test case giao dịch FX Hoán đổi"),
        ("M2-GTCG Govi Bond", "Test case giao dịch Trái phiếu Chính phủ"),
        ("M2-GTCG FI Bond & CCTG", "Test case giao dịch FI Bond và CCTG"),
        ("M3-MM Liên NH", "Test case giao dịch Liên ngân hàng"),
        ("M3-MM OMO", "Test case giao dịch thị trường mở"),
        ("M3-MM Repo KBNN", "Test case giao dịch Repo Kho bạc Nhà nước"),
        ("M4-Hạn mức", "Test case phê duyệt hạn mức liên ngân hàng"),
        ("M5-TTQT", "Test case thanh toán quốc tế"),
        ("Luồng đặc biệt", "Recall, Hủy, Clone giao dịch"),
        ("Phân quyền & Audit", "Phân quyền truy cập và kiểm toán"),
    ]
    for name, desc in sheets:
        ws.cell(row=r, column=2, value=name).font = Font(name="Arial", size=11, bold=True)
        ws.cell(row=r, column=3, value=desc).font = body_font
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="HƯỚNG DẪN ĐỌC").font = sub_font
    r += 1
    guides = [
        ("Cột 'Loại'", "Happy = Luồng chính thành công; Edge = Trường hợp biên; Negative = Kiểm tra lỗi/từ chối"),
        ("Cột 'Mức độ'", "Critical = Lỗi nghiêm trọng ảnh hưởng nghiệp vụ; High = Lỗi quan trọng; Medium = Lỗi trung bình; Low = Lỗi nhỏ"),
        ("Mã TC", "Định dạng: [Module]-[Loại]-[Số]. VD: FX-SP-001 = FX Spot test case #001"),
        ("Ghi chú", "Chứa lưu ý đặc biệt, tham chiếu BRD section, hoặc lỗi v1/v2 đã sửa"),
    ]
    for label, val in guides:
        ws.cell(row=r, column=2, value=label).font = Font(name="Arial", size=11, bold=True)
        ws.cell(row=r, column=3, value=val).font = body_font
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="COLOR LEGEND").font = sub_font
    r += 1
    legends = [
        ("Happy (Xanh lá)", HAPPY_FILL, HAPPY_FONT),
        ("Edge (Vàng cam)", EDGE_FILL, EDGE_FONT),
        ("Negative (Đỏ)", NEG_FILL, NEG_FONT),
    ]
    for label, fill, font in legends:
        c = ws.cell(row=r, column=2, value=label)
        c.fill = fill
        c.font = font
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="⚠️ TRỌNG TÂM KIỂM THỬ").font = sub_font
    r += 1
    focus = [
        "Công thức tính toán (Thành tiền FX, Lãi MM, Tổng giá trị GTCG, Hạn mức)",
        "Luồng phê duyệt đúng theo từng module (GTCG không qua QLRR/TTQT, OMO/Repo KBNN không qua QLRR/TTQT)",
        "Xác thực ngày tháng (Ngày Giá trị, Ngày đáo hạn, Kỳ hạn, năm nhuận)",
        "Ràng buộc dữ liệu (tồn kho GTCG khi bán, trường bắt buộc, số âm)",
        "Phân quyền xem dữ liệu theo role (QLRR chỉ xem MM Liên NH, KTTC chỉ từ bước hạch toán)",
        "Quy ước định dạng số (GTCG dùng VN format, FX/MM dùng quốc tế)",
        "KTTC 2 cấp (CV KTTC cấp 1 + LĐ KTTC cấp 2)",
        "Hạn mức v3: FX + MM dùng chung hạn mức; FI Bond dùng Giá thanh toán cho hạn mức không TSBĐ",
    ]
    for item in focus:
        ws.cell(row=r, column=2, value="•").font = body_font
        ws.cell(row=r, column=3, value=item).font = body_font
        r += 1


# ══════════════════════════════════════════════════════════════════════
# TEST CASE DATA
# ══════════════════════════════════════════════════════════════════════

def fx_spot_forward():
    """M1-FX Spot/Forward test cases."""
    tc = []
    n = [0]
    def add(code, group, desc, precond, steps, expected, typ, sev, note=""):
        n[0] += 1
        tc.append([n[0], code, group, desc, precond, steps, expected, typ, sev, note])

    # ── Happy Path ──
    add("FX-SP-001", "Tạo GD Spot", "Tạo giao dịch Spot Sell USD/VND thành công",
        "User CV K.NV đăng nhập; Đối tác MSB tồn tại trong master data",
        "1. Chọn Loại GD = SPOT\n2. Chiều = Sell\n3. Nhập KL = 1,000,000.00\n4. Loại tiền = USD\n5. Tỷ giá = 26,005.35\n6. Ngày GD = TODAY\n7. Ngày Giá trị = TODAY\n8. Chọn Pay code\n9. Nhấn Lưu",
        "GD tạo thành công, trạng thái Open.\nThành tiền = 1,000,000.00 × 26,005.35 = 26,005,350,000.00 VND",
        "Happy", "Critical", "Công thức: USD/VND → nhân")

    add("FX-SP-002", "Tạo GD Spot", "Tạo giao dịch Spot Buy EUR/USD thành công",
        "User CV K.NV; Đối tác ACB; Cặp tiền EUR/USD",
        "1. Chọn Loại GD = SPOT\n2. Chiều = Buy\n3. KL = 500,000.00\n4. Loại tiền = EUR\n5. Tỷ giá = 1.1550\n6. Nhập đầy đủ thông tin\n7. Lưu",
        "GD tạo thành công.\nThành tiền = 500,000.00 × 1.1550 = 577,500.00 USD\n(EUR/USD → .../USD → nhân)",
        "Happy", "Critical", "Công thức: .../USD → nhân")

    add("FX-SP-003", "Tạo GD Spot", "Tạo giao dịch Spot USD/JPY — công thức chia",
        "User CV K.NV; Cặp tiền USD/JPY",
        "1. SPOT, Sell\n2. KL = 2,000,000.00 JPY\n3. Tỷ giá = 150.25\n4. Lưu",
        "Thành tiền = 2,000,000.00 ÷ 150.25 = 13,311.15 USD\n(USD/JPY → USD/... → chia)",
        "Happy", "Critical", "Công thức: USD/... → chia")

    add("FX-SP-004", "Tạo GD Spot", "Tạo giao dịch Spot cặp tiền chéo EUR/GBP",
        "Cặp tiền EUR/GBP (không có USD)",
        "1. SPOT, Buy\n2. KL = 300,000.00 EUR\n3. Tỷ giá = 0.8650\n4. Lưu",
        "Thành tiền = 300,000.00 × 0.8650 = 259,500.0000 GBP\n(Cặp chéo → nhân, đơn vị = tiền quote)",
        "Happy", "High", "Công thức: Cặp chéo → nhân, kết quả = tiền quote")

    add("FX-SP-005", "Tạo GD Forward", "Tạo giao dịch Forward USD/VND thành công",
        "Đối tác BIDV; Loại GD = FORWARD",
        "1. Loại GD = FORWARD\n2. Chiều = Buy\n3. KL = 5,000,000.00\n4. Tỷ giá = 26,200.50\n5. Ngày Giá trị = TODAY + 30 ngày\n6. Lưu",
        "GD tạo thành công, trạng thái Open.\nThành tiền = 5,000,000.00 × 26,200.50 = 131,002,500,000.00 VND",
        "Happy", "High", "Forward cho phép Ngày Giá trị > TODAY")

    add("FX-SP-006", "Luồng duyệt", "Luồng phê duyệt FX Spot đầy đủ: CV → TP → GĐ → KTTC (2 cấp) → Hoàn thành",
        "GD FX Spot đã tạo, TTQT = Không",
        "1. CV submit GD\n2. TP K.NV duyệt cấp 1 → trạng thái Chờ K.NV duyệt\n3. GĐ TT KDV/QLV duyệt cấp 2 → Chờ hạch toán\n4. CV P.KTTC duyệt hạch toán cấp 1 → Chờ LĐ KTTC duyệt\n5. LĐ P.KTTC duyệt hạch toán cấp 2 → Hoàn thành",
        "GD chuyển qua đúng trình tự trạng thái.\nKhông qua QLRR.\nKết thúc = Hoàn thành.",
        "Happy", "Critical", "FX KHÔNG qua QLRR")

    add("FX-SP-007", "Luồng duyệt", "FX Spot có TTQT: CV → TP → GĐ → KTTC (2 cấp) → TTQT → Hoàn thành",
        "GD FX Spot có TTQT = Có",
        "1. CV submit\n2. TP duyệt\n3. GĐ duyệt\n4. CV KTTC duyệt cấp 1\n5. LĐ KTTC duyệt cấp 2\n6. BP.TTQT duyệt → Hoàn thành",
        "GD qua đúng 6 bước phê duyệt. Trạng thái cuối = Hoàn thành.",
        "Happy", "Critical", "")

    add("FX-SP-008", "Trường tự động", "Tên đối tác tự động hiển thị khi chọn Đối tác",
        "Master data có đối tác MSB = 'Ngân hàng TMCP Hàng hải Việt Nam'",
        "1. Tạo GD FX mới\n2. Chọn Đối tác = MSB",
        "Tên đối tác tự hiển thị: 'Ngân hàng TMCP Hàng hải Việt Nam'",
        "Happy", "Medium", "")

    add("FX-SP-009", "Trường tự động", "Cặp tiền tự động hiển thị theo Loại tiền giao dịch",
        "Loại tiền = EUR",
        "1. Chọn Loại tiền giao dịch = EUR",
        "Cặp tiền tự hiển thị: EUR/USD hoặc EUR/VND (tùy cấu hình)",
        "Happy", "Medium", "")

    add("FX-SP-010", "Trường tự động", "TTQT tự động = 'Có' khi Pay code đối tác là quốc tế",
        "Pay code đối tác = SWIFT code nước ngoài",
        "1. Chọn Pay code đối tác là tài khoản SWIFT nước ngoài",
        "Trường TTQT tự động = 'Có'",
        "Happy", "High", "")

    # ── Arithmetic/Formula ──
    add("FX-SP-011", "Công thức tính toán", "Thành tiền USD/VND với số lớn (tỷ VND)",
        "KL = 50,000,000.00 USD, Tỷ giá = 26,005.35",
        "1. Nhập KL = 50,000,000.00\n2. Tỷ giá = 26,005.35\n3. Kiểm tra Thành tiền",
        "Thành tiền = 50,000,000.00 × 26,005.35 = 1,300,267,500,000.00 VND\nKhông bị overflow, hiển thị đúng format",
        "Edge", "Critical", "Kiểm tra overflow số lớn")

    add("FX-SP-012", "Công thức tính toán", "Thành tiền với tỷ giá = 0 (division by zero cho USD/...)",
        "Cặp USD/JPY, Tỷ giá = 0",
        "1. Nhập KL = 1,000,000.00\n2. Tỷ giá = 0\n3. Lưu",
        "Hệ thống CHẶN, hiển thị lỗi 'Tỷ giá phải lớn hơn 0'.\nKhông được phép chia cho 0.",
        "Negative", "Critical", "Division by zero")

    add("FX-SP-013", "Công thức tính toán", "Tỷ giá decimal precision: USD/VND = 2 số, EUR/USD = 4 số",
        "Cặp USD/VND và EUR/USD",
        "1. Cặp USD/VND: nhập tỷ giá 26,005.3567 (4 số)\n2. Cặp EUR/USD: nhập tỷ giá 1.15 (2 số)",
        "USD/VND: Chỉ cho nhập 2 số sau dấu thập phân → 26,005.35\nEUR/USD: Cho phép 4 số → 1.1500",
        "Edge", "High", "BRD: USD/VND, USD/JPY, USD/KRW → 2 decimal; còn lại → 4")

    add("FX-SP-014", "Công thức tính toán", "Kiểm tra làm tròn Thành tiền với KL thập phân",
        "KL = 1,000,000.05 USD, Tỷ giá = 26,005.35",
        "1. Nhập KL có thập phân\n2. Kiểm tra kết quả",
        "Thành tiền = 1,000,000.05 × 26,005.35 = 26,005,351,300.18 VND (kiểm tra chính xác thập phân)",
        "Edge", "High", "Rounding errors")

    add("FX-SP-015", "Công thức tính toán", "Khối lượng giao dịch = 0",
        "KL = 0",
        "1. Nhập KL = 0\n2. Lưu",
        "Hệ thống CHẶN, hiển thị lỗi 'Khối lượng phải lớn hơn 0'",
        "Negative", "High", "")

    add("FX-SP-016", "Công thức tính toán", "Khối lượng giao dịch số âm",
        "KL = -1,000,000.00",
        "1. Nhập KL = -1,000,000.00\n2. Lưu",
        "Hệ thống CHẶN, không cho nhập số âm",
        "Negative", "High", "")

    # ── Date validation ──
    add("FX-SP-017", "Ngày tháng", "Ngày Giá trị = TODAY cho Spot",
        "Loại GD = SPOT",
        "1. Ngày GD = TODAY\n2. Ngày Giá trị = TODAY\n3. Lưu",
        "GD tạo thành công. Spot cho phép Ngày Giá trị = Ngày GD.",
        "Happy", "Medium", "")

    add("FX-SP-018", "Ngày tháng", "Ngày Giá trị < Ngày giao dịch (Spot)",
        "Loại GD = SPOT, Ngày GD = 03/04/2026, Ngày Giá trị = 01/04/2026",
        "1. Nhập Ngày GD = 03/04/2026\n2. Ngày Giá trị = 01/04/2026\n3. Lưu",
        "Hệ thống CHẶN: Ngày Giá trị không được trước Ngày giao dịch (đối với Spot)",
        "Negative", "High", "Forward có thể cho phép khác")

    add("FX-SP-019", "Ngày tháng", "Ngày Giá trị là ngày cuối tuần (Thứ 7/CN)",
        "Ngày Giá trị = 05/04/2026 (Chủ nhật)",
        "1. Chọn Ngày Giá trị = Chủ nhật\n2. Lưu",
        "Hệ thống CẢNH BÁO hoặc CHẶN ngày không phải ngày làm việc",
        "Edge", "Medium", "Tùy nghiệp vụ: cảnh báo hay block")

    add("FX-SP-020", "Ngày tháng", "Năm nhuận: Ngày Giá trị = 29/02/2028",
        "Năm 2028 là năm nhuận",
        "1. Nhập Ngày Giá trị = 29/02/2028\n2. Lưu",
        "GD tạo thành công. Hệ thống chấp nhận 29/02 trong năm nhuận.",
        "Edge", "Medium", "")

    add("FX-SP-021", "Ngày tháng", "Năm không nhuận: Ngày 29/02/2027 không tồn tại",
        "Năm 2027 không phải năm nhuận",
        "1. Nhập Ngày Giá trị = 29/02/2027",
        "Hệ thống CHẶN: Ngày không hợp lệ",
        "Negative", "Medium", "")

    add("FX-SP-022", "Ngày tháng", "Định dạng ngày dd/mm/yyyy (không nhầm mm/dd/yyyy)",
        "Ngày = 13/03/2026 (13 tháng 3)",
        "1. Nhập 13/03/2026\n2. Kiểm tra hiển thị",
        "Hiển thị đúng: 13/03/2026 = ngày 13 tháng 3. KHÔNG hiểu nhầm thành tháng 13.",
        "Edge", "High", "BRD quy ước dd/mm/yyyy")

    # ── Data constraints ──
    add("FX-SP-023", "Ràng buộc dữ liệu", "Tạo GD thiếu trường bắt buộc (Đối tác)",
        "Không chọn Đối tác",
        "1. Bỏ trống trường Đối tác\n2. Nhấn Lưu",
        "Hệ thống CHẶN, highlight trường Đối tác, hiển thị 'Vui lòng chọn Đối tác'",
        "Negative", "High", "")

    add("FX-SP-024", "Ràng buộc dữ liệu", "Tạo GD thiếu trường Tỷ giá",
        "Không nhập Tỷ giá",
        "1. Điền đầy đủ trừ Tỷ giá\n2. Nhấn Lưu",
        "Hệ thống CHẶN, hiển thị lỗi bắt buộc nhập Tỷ giá",
        "Negative", "High", "")

    add("FX-SP-025", "Ràng buộc dữ liệu", "Số Ticket là không bắt buộc",
        "Bỏ trống Số Ticket",
        "1. Tạo GD không nhập Số Ticket\n2. Lưu",
        "GD tạo thành công. Số Ticket = trống.",
        "Happy", "Low", "")

    add("FX-SP-026", "Ràng buộc dữ liệu", "Tỷ giá âm",
        "Nhập Tỷ giá = -26,005.35",
        "1. Nhập Tỷ giá = -26,005.35\n2. Lưu",
        "Hệ thống CHẶN, không cho nhập tỷ giá âm",
        "Negative", "High", "")

    add("FX-SP-027", "Định dạng số", "Format số tiền FX dùng quốc tế (dấu , ngăn nghìn, . thập phân)",
        "KL = 10,000,000.05",
        "1. Nhập KL\n2. Kiểm tra hiển thị",
        "Hiển thị: 10,000,000.05 (dấu , ngăn nghìn, dấu . thập phân)\nKHÔNG hiển thị kiểu VN: 10.000.000,05",
        "Edge", "High", "BRD v3: FX dùng format quốc tế")

    add("FX-SP-028", "Upload file", "Upload file đính kèm ticket",
        "Có file ticket_925508281.pdf",
        "1. Tạo GD\n2. Upload file ticket\n3. Lưu",
        "File đính kèm thành công, có thể tải lại xem",
        "Happy", "Low", "1 file = 1 GD")

    add("FX-SP-029", "Danh sách GD", "Lọc giao dịch theo trạng thái",
        "Có nhiều GD FX ở các trạng thái khác nhau",
        "1. Mở danh sách GD FX\n2. Lọc theo trạng thái = 'Chờ hạch toán'",
        "Chỉ hiển thị GD có trạng thái 'Chờ hạch toán'",
        "Happy", "Medium", "")

    add("FX-SP-030", "Danh sách GD", "GD hủy mặc định ẩn, filter hiện",
        "Có GD FX đã hủy",
        "1. Mở danh sách → GD hủy không hiển thị\n2. Bật filter 'Hiển thị GD đã hủy'\n3. GD hủy xuất hiện",
        "Mặc định ẩn GD hủy. Khi bật filter → hiển thị.",
        "Happy", "Medium", "")

    add("FX-SP-031", "Danh sách GD", "Tìm kiếm theo Số Ticket",
        "GD có Số Ticket = 925508281",
        "1. Nhập 925508281 vào ô tìm kiếm\n2. Nhấn tìm",
        "Hiển thị GD có Số Ticket = 925508281",
        "Happy", "Medium", "")

    add("FX-SP-032", "Luồng duyệt", "TP K.NV trả lại CV sửa → GD quay về Open",
        "GD ở trạng thái Open, CV đã submit",
        "1. CV submit GD\n2. TP K.NV chọn 'Không đồng ý'",
        "GD quay về trạng thái Open. CV có thể chỉnh sửa.",
        "Happy", "High", "")

    add("FX-SP-033", "Luồng duyệt", "GĐ từ chối → popup xác nhận → trạng thái Từ chối",
        "GD ở trạng thái Chờ K.NV duyệt",
        "1. GĐ nhấn 'Không đồng ý'\n2. Popup hiện: nhập lý do + nút Xác nhận/Hủy\n3. Nhập lý do, nhấn Xác nhận",
        "Popup xác nhận hiển thị. Sau xác nhận → GD = 'Từ chối'. Thông báo cho CV.",
        "Happy", "Critical", "Popup chống click nhầm")

    add("FX-SP-034", "Luồng duyệt", "GĐ từ chối → nhấn Hủy trên popup → GD giữ nguyên",
        "GD ở trạng thái Chờ K.NV duyệt",
        "1. GĐ nhấn 'Không đồng ý'\n2. Popup hiện\n3. Nhấn 'Hủy' trên popup",
        "Popup đóng. GD giữ nguyên trạng thái 'Chờ K.NV duyệt'.",
        "Edge", "High", "")

    add("FX-SP-035", "Luồng duyệt", "CV KTTC từ chối hạch toán cấp 1 → Hủy giao dịch",
        "GD ở trạng thái Chờ hạch toán",
        "1. CV KTTC nhấn Từ chối\n2. Popup xác nhận → nhập lý do → Xác nhận",
        "GD chuyển = 'Hủy giao dịch'. Thông báo cho CV K.NV.",
        "Happy", "Critical", "KTTC 2 cấp - cấp 1 từ chối")

    add("FX-SP-036", "Luồng duyệt", "LĐ KTTC từ chối hạch toán cấp 2 → Hủy giao dịch",
        "GD ở trạng thái Chờ LĐ KTTC duyệt (CV KTTC đã duyệt cấp 1)",
        "1. LĐ KTTC nhấn Từ chối\n2. Popup xác nhận → nhập lý do → Xác nhận",
        "GD chuyển = 'Hủy giao dịch'. Thông báo cho CV K.NV.",
        "Happy", "Critical", "KTTC 2 cấp - cấp 2 từ chối")

    add("FX-SP-037", "Workflow", "CV chỉnh sửa GD sau khi TP đã duyệt → BLOCK",
        "GD ở trạng thái 'Chờ K.NV duyệt' (TP đã duyệt)",
        "1. CV mở GD đang Chờ K.NV duyệt\n2. Cố chỉnh sửa trường KL hoặc Tỷ giá",
        "Hệ thống CHẶN: Không cho phép chỉnh sửa khi GD đã qua bước TP duyệt.",
        "Negative", "Critical", "BRD: Sau TP duyệt → CV không thể sửa")

    add("FX-SP-038", "Ghi chú", "Nhập ghi chú dạng text tự do",
        "GD FX mới",
        "1. Nhập Ghi chú = 'Giao dịch theo chỉ đạo TGĐ ngày 03/04'\n2. Lưu",
        "Ghi chú lưu thành công, hiển thị đúng nội dung",
        "Happy", "Low", "")

    return tc


def fx_swap():
    """M1-FX Swap test cases."""
    tc = []
    n = [0]
    def add(code, group, desc, precond, steps, expected, typ, sev, note=""):
        n[0] += 1
        tc.append([n[0], code, group, desc, precond, steps, expected, typ, sev, note])

    add("FX-SW-001", "Tạo GD Swap", "Tạo Swap Sell-Buy USD/VND thành công",
        "CV K.NV; Đối tác BIDV",
        "1. Chiều = Sell - Buy\n2. KL = 10,000,000.00 USD\n3. Cặp USD/VND\n4. Tỷ giá chân 1 = 26,005.35\n5. Tỷ giá chân 2 = 26,105.00\n6. Ngày Giá trị chân 1 = TODAY\n7. Ngày Giá trị chân 2 = TODAY + 90 ngày\n8. Pay code chân 1 & chân 2\n9. Lưu",
        "GD tạo thành công.\nThành tiền chân 1 = 10M × 26,005.35 = 260,053,500,000 VND\nThành tiền chân 2 = 10M × 26,105.00 = 261,050,000,000 VND",
        "Happy", "Critical", "Chân 2 dùng Tỷ giá chân 2 (sửa lỗi BRD v1)")

    add("FX-SW-002", "Công thức", "Swap chân 2 PHẢI dùng Tỷ giá chân 2 (không phải chân 1)",
        "Tỷ giá chân 1 = 26,005.35; Tỷ giá chân 2 = 26,105.00; KL = 10,000,000.00",
        "1. Tạo Swap USD/VND\n2. Kiểm tra Thành tiền chân 2",
        "Thành tiền chân 2 = 10,000,000.00 × 26,105.00 = 261,050,000,000.00 VND\nNẾU kết quả = 260,053,500,000 → BUG (đang dùng tỷ giá chân 1)",
        "Happy", "Critical", "⚠️ BRD v1 có lỗi typo dùng tỷ giá chân 1 cho chân 2")

    add("FX-SW-003", "Công thức", "Swap USD/JPY — chân 1 chia, chân 2 chia (với tỷ giá riêng)",
        "Cặp USD/JPY; TG chân 1 = 150.25; TG chân 2 = 151.00",
        "1. KL = 5,000,000.00 JPY\n2. Kiểm tra cả 2 chân",
        "Chân 1 = 5,000,000 ÷ 150.25 = 33,277.87 USD\nChân 2 = 5,000,000 ÷ 151.00 = 33,112.58 USD",
        "Happy", "Critical", "USD/... → chia")

    add("FX-SW-004", "Công thức", "Swap AUD/USD — chân 1 nhân, chân 2 nhân",
        "Cặp AUD/USD; TG chân 1 = 0.6750; TG chân 2 = 0.6800",
        "1. KL = 2,000,000.00 AUD\n2. Kiểm tra",
        "Chân 1 = 2,000,000 × 0.6750 = 1,350,000.00 USD\nChân 2 = 2,000,000 × 0.6800 = 1,360,000.00 USD",
        "Happy", "Critical", ".../USD → nhân")

    add("FX-SW-005", "Công thức", "Swap cặp chéo EUR/GBP",
        "Cặp EUR/GBP; TG chân 1 = 0.8650; TG chân 2 = 0.8700",
        "1. KL = 1,000,000.00 EUR\n2. Kiểm tra",
        "Chân 1 = 1,000,000 × 0.8650 = 865,000.0000 GBP\nChân 2 = 1,000,000 × 0.8700 = 870,000.0000 GBP",
        "Happy", "High", "Cặp chéo → nhân, kết quả = tiền quote")

    add("FX-SW-006", "Ngày tháng", "Ngày Giá trị chân 2 phải SAU chân 1",
        "Ngày Giá trị chân 1 = 03/04/2026; Ngày Giá trị chân 2 = 03/07/2026",
        "1. Nhập ngày chân 2 sau chân 1\n2. Lưu",
        "GD tạo thành công. Chân 2 > chân 1.",
        "Happy", "High", "")

    add("FX-SW-007", "Ngày tháng", "Ngày Giá trị chân 2 = chân 1 → BLOCK",
        "Ngày Giá trị chân 1 = 03/04/2026; chân 2 = 03/04/2026",
        "1. Nhập chân 2 = chân 1\n2. Lưu",
        "Hệ thống CHẶN: Ngày Giá trị chân 2 phải sau chân 1",
        "Negative", "High", "")

    add("FX-SW-008", "Ngày tháng", "Ngày Giá trị chân 2 TRƯỚC chân 1 → BLOCK",
        "Ngày chân 1 = 03/07/2026; chân 2 = 03/04/2026",
        "1. Nhập chân 2 trước chân 1\n2. Lưu",
        "Hệ thống CHẶN: Ngày Giá trị chân 2 phải sau chân 1",
        "Negative", "Critical", "")

    add("FX-SW-009", "Ngày tháng", "Swap qua năm (chân 1 = 2026, chân 2 = 2027)",
        "Ngày chân 1 = 15/12/2026; chân 2 = 15/03/2027",
        "1. Tạo Swap qua năm\n2. Lưu",
        "GD tạo thành công. Hệ thống xử lý đúng giao dịch qua năm.",
        "Edge", "Medium", "")

    add("FX-SW-010", "Tỷ giá", "Tỷ giá chân 1 = 0 → division by zero (USD/...)",
        "Cặp USD/JPY; TG chân 1 = 0",
        "1. Nhập TG chân 1 = 0\n2. Lưu",
        "Hệ thống CHẶN: Tỷ giá phải > 0",
        "Negative", "Critical", "Division by zero cho công thức chia")

    add("FX-SW-011", "Tỷ giá", "Tỷ giá chân 2 = 0",
        "Cặp USD/JPY; TG chân 2 = 0",
        "1. TG chân 2 = 0\n2. Lưu",
        "Hệ thống CHẶN: Tỷ giá chân 2 phải > 0",
        "Negative", "Critical", "")

    add("FX-SW-012", "Luồng duyệt", "Swap đi qua đúng luồng FX (không qua QLRR)",
        "GD Swap tạo thành công",
        "1. CV submit\n2. TP duyệt\n3. GĐ duyệt\n4. KTTC cấp 1 duyệt\n5. KTTC cấp 2 duyệt\n6. TTQT (nếu có)",
        "Luồng Swap giống Spot/Forward.\nKHÔNG qua QLRR.",
        "Happy", "Critical", "")

    add("FX-SW-013", "TTQT", "Swap chân 1 có TTQT, chân 2 có TTQT",
        "Pay code đối tác chân 1 & chân 2 đều SWIFT quốc tế",
        "1. Tạo Swap\n2. Kiểm tra TTQT chân 1 & chân 2",
        "TTQT chân 1 = Có; TTQT chân 2 = Có.\nChân 1 và chân 2 xuất hiện trên TTQT theo đúng ngày thực hiện.",
        "Happy", "High", "")

    add("FX-SW-014", "TTQT", "Ticket Swap: chân 1 suffix 'A', chân 2 suffix 'B'",
        "Số Ticket gốc = 12345",
        "1. Tạo Swap có Số Ticket = 12345\n2. Kiểm tra trên màn hình TTQT",
        "TTQT chân 1: Ticket = 12345A\nTTQT chân 2: Ticket = 12345B",
        "Happy", "High", "BRD: suffix A/B")

    add("FX-SW-015", "Ràng buộc", "Thiếu Pay code chân 1",
        "Không nhập Pay code KLB chân 1",
        "1. Bỏ trống Pay code KLB chân 1\n2. Lưu",
        "Hệ thống CHẶN: Pay code KLB chân 1 bắt buộc",
        "Negative", "High", "")

    add("FX-SW-016", "Ràng buộc", "Thiếu Pay code chân 2",
        "Không nhập Pay code đối tác chân 2",
        "1. Bỏ trống\n2. Lưu",
        "Hệ thống CHẶN: Pay code đối tác chân 2 bắt buộc",
        "Negative", "High", "")

    add("FX-SW-017", "Công thức", "Swap với số lượng cực lớn (100M USD × 26,000)",
        "KL = 100,000,000.00 USD; TG = 26,000.00",
        "1. Tạo Swap\n2. Kiểm tra Thành tiền",
        "Thành tiền = 2,600,000,000,000.00 VND (2.6 nghìn tỷ). Không overflow.",
        "Edge", "High", "Kiểm tra overflow")

    add("FX-SW-018", "Tỷ giá", "Tỷ giá chân 1 = tỷ giá chân 2 (Swap flat)",
        "TG chân 1 = TG chân 2 = 26,005.35",
        "1. Nhập TG giống nhau\n2. Lưu",
        "GD tạo thành công. Cho phép TG chân 1 = chân 2 (Swap flat).",
        "Edge", "Low", "")

    add("FX-SW-019", "Ngày tháng", "Swap 29/02 năm nhuận (chân 2)",
        "Chân 2 = 29/02/2028",
        "1. Nhập ngày chân 2 = 29/02/2028\n2. Lưu",
        "GD tạo thành công. Ngày hợp lệ trong năm nhuận.",
        "Edge", "Medium", "")

    add("FX-SW-020", "Workflow", "CV sửa Swap sau TP duyệt → BLOCK",
        "Swap ở trạng thái Chờ K.NV duyệt",
        "1. CV mở Swap\n2. Cố sửa TG chân 2",
        "Hệ thống CHẶN: Không cho sửa sau TP duyệt.",
        "Negative", "Critical", "")

    add("FX-SW-021", "Chiều GD", "Swap Buy-Sell thành công",
        "Chiều = Buy - Sell",
        "1. Chọn Chiều = Buy - Sell\n2. Điền đầy đủ\n3. Lưu",
        "GD tạo thành công. Chiều GD = Buy - Sell.",
        "Happy", "High", "")

    add("FX-SW-022", "Tỷ giá", "Tỷ giá âm cho chân 1",
        "TG chân 1 = -26,005.35",
        "1. Nhập TG chân 1 âm\n2. Lưu",
        "Hệ thống CHẶN: Tỷ giá phải > 0",
        "Negative", "High", "")

    add("FX-SW-023", "Tỷ giá", "Tỷ giá decimal precision cho Swap",
        "USD/VND: TG chân 1 = 26,005.3567 (4 chữ số)",
        "1. Nhập TG chân 1 với 4 decimal cho USD/VND",
        "Hệ thống chỉ cho 2 decimal: 26,005.35. Cắt/làm tròn.",
        "Edge", "High", "")

    add("FX-SW-024", "File", "Upload file đính kèm cho Swap",
        "File hợp đồng swap",
        "1. Tạo Swap\n2. Upload file\n3. Lưu",
        "File lưu thành công",
        "Happy", "Low", "")

    add("FX-SW-025", "Tạo GD Swap", "Tạo Swap không nhập Số Ticket (không bắt buộc)",
        "Bỏ trống Số Ticket",
        "1. Tạo Swap không nhập Ticket\n2. Lưu",
        "GD tạo thành công. Số Ticket = trống.",
        "Happy", "Low", "")

    return tc


def gtcg_govi():
    """M2-GTCG Govi Bond test cases."""
    tc = []
    n = [0]
    def add(code, group, desc, precond, steps, expected, typ, sev, note=""):
        n[0] += 1
        tc.append([n[0], code, group, desc, precond, steps, expected, typ, sev, note])

    add("GT-GV-001", "Tạo GD mua", "Tạo giao dịch mua Govi Bond Outright thành công",
        "CV K.NV; Mã TP = TD2135068 tồn tại trong danh mục",
        "1. Chiều = Bên mua\n2. Loại GD = Outright\n3. Mã TP = TD2135068\n4. SL = 1.000.000\n5. Giá thanh toán = 1.206.307.366\n6. Hạch toán = HTM\n7. Lưu",
        "GD tạo thành công.\nTổng giá trị = 1.000.000 × 1.206.307.366 = 1.206.307.366.000.000\nFormat VN: dấu . ngăn nghìn",
        "Happy", "Critical", "GTCG dùng format VN (. ngăn nghìn)")

    add("GT-GV-002", "Công thức", "Tổng giá trị = Số lượng × Giá thanh toán",
        "SL = 500.000; Giá TT = 120.160",
        "1. Nhập SL và Giá TT\n2. Kiểm tra Tổng giá trị",
        "Tổng giá trị = 500.000 × 120.160 = 60.080.000.000",
        "Happy", "Critical", "")

    add("GT-GV-003", "Công thức", "Kỳ hạn còn lại = Ngày đáo hạn - Ngày thanh toán",
        "Ngày đáo hạn = 12/03/2035; Ngày TT = 13/03/2026",
        "1. Chọn Mã TP có đáo hạn 12/03/2035\n2. Ngày TT = 13/03/2026\n3. Kiểm tra Kỳ hạn còn lại",
        "Kỳ hạn còn lại = 3.286 ngày (tính từ 13/03/2026 → 12/03/2035)",
        "Happy", "High", "")

    add("GT-GV-004", "Tồn kho", "Bán vượt tồn kho → BLOCK cứng",
        "Mã TP TD2135068 tồn kho = 100.000; Chiều = Bên bán",
        "1. Chiều = Bên bán\n2. SL = 150.000 (> tồn kho 100.000)\n3. Lưu",
        "Hệ thống CHẶN CỨNG: 'Số lượng bán (150.000) vượt tồn kho hiện có (100.000)'. Không cho tạo GD.",
        "Negative", "Critical", "BRD: Block cứng khi bán vượt tồn kho")

    add("GT-GV-005", "Tồn kho", "Bán = đúng tồn kho → cho phép",
        "Tồn kho = 100.000; SL bán = 100.000",
        "1. SL bán = 100.000\n2. Lưu",
        "GD tạo thành công. Bán hết tồn kho cho phép.",
        "Edge", "High", "")

    add("GT-GV-006", "Tồn kho", "Bán < tồn kho → cho phép",
        "Tồn kho = 100.000; SL bán = 50.000",
        "1. SL bán = 50.000\n2. Lưu",
        "GD tạo thành công.",
        "Happy", "High", "")

    add("GT-GV-007", "Trường tự động", "Lãi suất coupon, Ngày phát hành, Ngày đáo hạn, Mệnh giá tự động",
        "Mã TP = TD2135068 trong danh mục",
        "1. Chọn Mã TP = TD2135068",
        "Tự động điền: Lãi suất coupon, Ngày phát hành, Ngày đáo hạn, Mệnh giá từ danh mục",
        "Happy", "High", "")

    add("GT-GV-008", "Luồng duyệt", "GTCG KHÔNG qua QLRR, KHÔNG qua TTQT",
        "GD GTCG Govi Bond tạo thành công",
        "1. CV submit\n2. TP duyệt\n3. GĐ duyệt\n4. CV KTTC cấp 1\n5. LĐ KTTC cấp 2\n→ Hoàn thành",
        "GD hoàn thành qua 5 bước.\nKHÔNG có bước QLRR.\nKHÔNG có bước TTQT.",
        "Happy", "Critical", "BRD v3: GTCG không qua QLRR, không qua TTQT")

    add("GT-GV-009", "Luồng duyệt", "Verify GTCG không hiển thị trạng thái 'Chờ QLRR'",
        "GD GTCG đã tạo",
        "1. Theo dõi GD từ Open → Hoàn thành\n2. Kiểm tra không có trạng thái 'Chờ QLRR'",
        "Danh sách trạng thái GTCG KHÔNG bao gồm 'Chờ QLRR' và 'Chờ TTQT'",
        "Negative", "Critical", "Nếu xuất hiện Chờ QLRR → BUG")

    add("GT-GV-010", "Hạch toán", "Trường Hạch toán (HTM/AFS/HFT) chỉ hiện khi mua",
        "Chiều = Bên mua",
        "1. Chọn Chiều = Bên mua → trường Hạch toán hiển thị\n2. Đổi sang Bên bán → trường Hạch toán ẩn",
        "Hạch toán hiển thị khi mua, ẩn khi bán.",
        "Happy", "High", "")

    add("GT-GV-011", "Hạch toán", "Bên bán không cần chọn Hạch toán",
        "Chiều = Bên bán",
        "1. Chiều = Bên bán\n2. Kiểm tra trường Hạch toán",
        "Trường Hạch toán không hiển thị. Không bắt buộc.",
        "Happy", "Medium", "")

    add("GT-GV-012", "Định dạng", "GTCG dùng VN format (dấu . ngăn nghìn, dấu , thập phân)",
        "SL = 1.000.000; Giá = 120.160",
        "1. Nhập SL = 1000000\n2. Kiểm tra hiển thị",
        "Hiển thị: 1.000.000 (dấu . ngăn nghìn)\nLãi suất: 2,51% (dấu , thập phân)\nKHÔNG dùng format quốc tế",
        "Edge", "High", "BRD v3: GTCG dùng VN format, khác FX/MM")

    add("GT-GV-013", "Ràng buộc", "Số lượng phải là số nguyên (không thập phân)",
        "SL = 1.000.000,5 (có thập phân)",
        "1. Nhập SL = 1000000.5\n2. Lưu",
        "Hệ thống CHẶN: Số lượng phải là số nguyên, không cho phép thập phân",
        "Negative", "High", "")

    add("GT-GV-014", "Ràng buộc", "Số lượng = 0 → BLOCK",
        "SL = 0",
        "1. Nhập SL = 0\n2. Lưu",
        "Hệ thống CHẶN: Số lượng phải > 0",
        "Negative", "High", "")

    add("GT-GV-015", "Ràng buộc", "Số lượng âm → BLOCK",
        "SL = -100",
        "1. Nhập SL = -100\n2. Lưu",
        "Hệ thống CHẶN: Không cho phép số lượng âm",
        "Negative", "High", "")

    add("GT-GV-016", "Ngày tháng", "Kỳ hạn còn lại = 0 khi Ngày TT = Ngày đáo hạn",
        "Ngày đáo hạn = 13/03/2026; Ngày TT = 13/03/2026",
        "1. Chọn TP có đáo hạn = ngày TT\n2. Kiểm tra Kỳ hạn",
        "Kỳ hạn còn lại = 0 ngày. Hệ thống hiển thị đúng.",
        "Edge", "Medium", "")

    add("GT-GV-017", "Ngày tháng", "Kỳ hạn còn lại âm (Ngày TT > Ngày đáo hạn)",
        "Ngày đáo hạn = 12/03/2026; Ngày TT = 15/03/2026",
        "1. Chọn TP đã quá đáo hạn\n2. Kiểm tra",
        "Hệ thống CẢNH BÁO: Trái phiếu đã qua ngày đáo hạn. Kỳ hạn còn lại = -3 ngày.",
        "Edge", "High", "Trái phiếu hết hạn")

    add("GT-GV-018", "Trường tìm kiếm", "Dropdown Mã TP có tìm kiếm gợi ý",
        "Danh mục có nhiều TP bắt đầu bằng 'TD'",
        "1. Nhập 'TD' vào trường Mã TP\n2. Kiểm tra gợi ý",
        "Hệ thống hiển thị danh sách gợi ý các TP bắt đầu bằng 'TD'",
        "Happy", "Medium", "")

    add("GT-GV-019", "Xác nhận GD", "Chọn Xác nhận = 'Khác' → cho phép nhập text",
        "Xác nhận GD = Khác",
        "1. Chọn 'Khác'\n2. Nhập text 'Bloomberg chat'",
        "Cho phép nhập text tự do khi chọn 'Khác'",
        "Happy", "Low", "")

    add("GT-GV-020", "Loại GD", "Chọn Loại GD = 'Other' → cho phép nhập tay",
        "Loại GD = Other",
        "1. Chọn Other\n2. Nhập tay 'Mua lại'",
        "Cho phép nhập text khi chọn Other",
        "Happy", "Low", "")

    add("GT-GV-021", "Ghi chú", "Bên bán: Ghi chú ngày mua và giá mua cho P.KTTC",
        "Chiều = Bên bán",
        "1. Nhập Ghi chú = 'Ngày mua: 12/03/2026, Giá mua: 102.260 VND/TP'\n2. Lưu",
        "Ghi chú hiển thị đúng cho P.KTTC tham khảo",
        "Happy", "Medium", "BRD yêu cầu ghi chú giá mua khi bán")

    add("GT-GV-022", "Mã GD", "Mã giao dịch Govi Bond tự sinh format Gxxxxxxxxxx",
        "Tạo GD Govi Bond thành công",
        "1. Tạo GD Govi Bond\n2. Kiểm tra mã GD trên danh sách",
        "Mã GD bắt đầu bằng 'G', VD: G0000000001",
        "Happy", "Medium", "")

    add("GT-GV-023", "Ràng buộc", "Lãi suất chiết khấu: tối đa 4 chữ số thập phân",
        "LS chiết khấu = 2,56789% (5 decimal)",
        "1. Nhập LS chiết khấu = 2,56789%",
        "Hệ thống chỉ chấp nhận 4 decimal: 2,5678% hoặc CHẶN",
        "Edge", "Medium", "")

    add("GT-GV-024", "Đối tác", "Cho phép nhập bổ sung đối tác mới",
        "Đối tác mới chưa có trong dropdown",
        "1. Tại trường Đối tác, nhập tên đối tác mới\n2. Kiểm tra",
        "Cho phép nhập đối tác mới (BRD: 'Cho phép nhập bổ sung đối tác mới')",
        "Happy", "Medium", "")

    add("GT-GV-025", "Luồng duyệt", "GĐ từ chối → Popup xác nhận → trạng thái Từ chối",
        "GD Govi Bond Chờ K.NV duyệt",
        "1. GĐ nhấn Từ chối\n2. Popup hiện\n3. Nhập lý do\n4. Xác nhận",
        "GD = Từ chối. Notify CV K.NV.",
        "Happy", "Critical", "")

    add("GT-GV-026", "Luồng duyệt", "CV KTTC cấp 1 từ chối → Hủy giao dịch",
        "GD ở Chờ hạch toán",
        "1. CV KTTC từ chối\n2. Popup → lý do → Xác nhận",
        "GD = Hủy giao dịch. Notify CV.",
        "Happy", "Critical", "")

    add("GT-GV-027", "Luồng duyệt", "LĐ KTTC cấp 2 từ chối → Hủy giao dịch",
        "GD ở Chờ LĐ KTTC duyệt",
        "1. LĐ KTTC từ chối\n2. Popup → lý do → Xác nhận",
        "GD = Hủy giao dịch. Notify CV.",
        "Happy", "Critical", "")

    add("GT-GV-028", "Workflow", "CV sửa GD Govi Bond sau TP duyệt → BLOCK",
        "GD ở Chờ K.NV duyệt",
        "1. CV mở GD\n2. Cố sửa SL",
        "Hệ thống CHẶN: Không cho sửa sau TP duyệt.",
        "Negative", "Critical", "")

    add("GT-GV-029", "Tổng giá trị", "Tổng giá trị với SL lớn (overflow check)",
        "SL = 10.000.000; Giá TT = 1.206.307.366",
        "1. Nhập SL = 10.000.000\n2. Kiểm tra Tổng giá trị",
        "Tổng = 12.063.073.660.000.000 (hơn 12 nghìn tỷ). Không overflow.",
        "Edge", "High", "")

    add("GT-GV-030", "Repo GD", "Tạo GD Govi Bond Repo thành công",
        "Loại GD = Repo",
        "1. Chọn Repo\n2. Điền đầy đủ\n3. Lưu",
        "GD Repo tạo thành công",
        "Happy", "High", "")

    return tc


def gtcg_fi_cctg():
    """M2-GTCG FI Bond & CCTG test cases."""
    tc = []
    n = [0]
    def add(code, group, desc, precond, steps, expected, typ, sev, note=""):
        n[0] += 1
        tc.append([n[0], code, group, desc, precond, steps, expected, typ, sev, note])

    add("GT-FI-001", "Tạo GD FI Bond", "Tạo giao dịch FI Bond mua Outright thành công",
        "CV K.NV; Loại GTCG = FI Bond",
        "1. Loại GTCG = FI Bond\n2. Chiều = Bên mua\n3. Đối tác = MSB\n4. Loại GD = Outright\n5. Nhập đầy đủ SL, Mệnh giá, Giá sạch, Giá TT\n6. Hạch toán = AFS\n7. Lưu",
        "GD tạo thành công. Format VN (dấu . ngăn nghìn).",
        "Happy", "Critical", "")

    add("GT-FI-002", "Tạo GD CCTG", "Tạo giao dịch CCTG mua thành công",
        "Loại GTCG = CCTG",
        "1. Loại GTCG = CCTG\n2. Chiều = Bên mua\n3. Điền đầy đủ\n4. Lưu",
        "GD CCTG tạo thành công. Mã GD: Fxxxxxxxxxx (FI Bond) hoặc format tương ứng.",
        "Happy", "Critical", "")

    add("GT-FI-003", "Công thức", "Tổng giá trị = Số lượng × Giá thanh toán (FI Bond)",
        "SL = 200.000; Giá TT = 150.250",
        "1. Nhập SL và Giá TT\n2. Kiểm tra",
        "Tổng giá trị = 200.000 × 150.250 = 30.050.000.000",
        "Happy", "Critical", "")

    add("GT-FI-004", "Công thức", "Kỳ hạn còn lại FI Bond = Ngày đáo hạn - Ngày TT",
        "Ngày đáo hạn = 11/03/2027; Ngày TT = 14/04/2026",
        "1. Kiểm tra Kỳ hạn còn lại",
        "Kỳ hạn = 331 ngày",
        "Happy", "High", "")

    add("GT-FI-005", "Tồn kho", "FI Bond bán vượt tồn kho → BLOCK",
        "Tồn kho Mã GTCG = 50.000; SL bán = 60.000",
        "1. Chiều = Bên bán\n2. SL = 60.000\n3. Lưu",
        "Hệ thống CHẶN CỨNG: Không cho bán vượt tồn kho",
        "Negative", "Critical", "")

    add("GT-FI-006", "Tồn kho", "CCTG bán vượt tồn kho → BLOCK",
        "Tồn kho CCTG = 30.000; SL bán = 40.000",
        "1. Chiều = Bên bán\n2. SL = 40.000\n3. Lưu",
        "Hệ thống CHẶN CỨNG",
        "Negative", "Critical", "")

    add("GT-FI-007", "Luồng duyệt", "FI Bond KHÔNG qua QLRR, KHÔNG qua TTQT",
        "GD FI Bond",
        "1. Theo dõi luồng: CV → TP → GĐ → KTTC (2 cấp) → Hoàn thành",
        "KHÔNG có bước QLRR. KHÔNG có bước TTQT.",
        "Happy", "Critical", "")

    add("GT-FI-008", "Hạn mức", "FI Bond: Hạn mức không TSBĐ dùng Giá thanh toán (không phải Số tiền gốc)",
        "FI Bond đã hoàn thành; Đối tác = MSB",
        "1. Kiểm tra Module 4 Hạn mức\n2. Hạn mức đã sử dụng (không TSBĐ)",
        "Hạn mức đã SD tính theo 'Giá thanh toán' GD FI Bond, KHÔNG phải 'Số tiền gốc'",
        "Happy", "Critical", "BRD v3: FI Bond dùng Giá thanh toán cho hạn mức ko TSBĐ")

    add("GT-FI-009", "Định dạng", "FI Bond dùng VN format giống Govi Bond",
        "Nhập SL = 1000000",
        "1. Kiểm tra hiển thị",
        "Hiển thị: 1.000.000 (VN format). Lãi suất: 2,5000% (dấu , thập phân)",
        "Edge", "High", "")

    add("GT-FI-010", "Ràng buộc", "Mã GTCG không bắt buộc cho FI Bond",
        "Bỏ trống Mã GTCG",
        "1. Tạo FI Bond, bỏ trống Mã GTCG\n2. Lưu",
        "GD tạo thành công. Mã GTCG là không bắt buộc.",
        "Happy", "Medium", "BRD: Mã GTCG = Không bắt buộc")

    add("GT-FI-011", "Ràng buộc", "Tổ chức phát hành bắt buộc",
        "Bỏ trống TCPH",
        "1. Bỏ trống Tổ chức phát hành\n2. Lưu",
        "Hệ thống CHẶN: TCPH bắt buộc",
        "Negative", "High", "")

    add("GT-FI-012", "Ràng buộc", "Lãi suất coupon: 4 chữ số thập phân",
        "LS = 2,50001% (5 decimal)",
        "1. Nhập LS coupon = 2,50001%",
        "Hệ thống chỉ cho 4 decimal: 2,5000%",
        "Edge", "Medium", "")

    add("GT-FI-013", "Ngày tháng", "Ngày đáo hạn FI Bond/CCTG — nhập thủ công (bắt buộc)",
        "FI Bond (khác Govi Bond — Govi lấy từ danh mục)",
        "1. Nhập Ngày đáo hạn = 11/03/2027\n2. Lưu",
        "Ngày đáo hạn lưu thành công. FI Bond cho nhập tay.",
        "Happy", "Medium", "Govi: tự động từ danh mục; FI: nhập tay")

    add("GT-FI-014", "Hạch toán", "FI Bond mua → trường Hạch toán hiển thị",
        "Chiều = Bên mua",
        "1. Kiểm tra trường Hạch toán",
        "Hiển thị dropdown HTM/AFS/HFT khi mua",
        "Happy", "High", "")

    add("GT-FI-015", "Hạch toán", "FI Bond bán → trường Hạch toán ẩn",
        "Chiều = Bên bán",
        "1. Kiểm tra trường Hạch toán",
        "Trường Hạch toán không hiển thị",
        "Happy", "Medium", "")

    add("GT-FI-016", "Ràng buộc", "SL phải số nguyên (FI Bond)",
        "SL = 100.500,5 (có thập phân)",
        "1. Nhập SL có thập phân",
        "Hệ thống CHẶN: Số lượng phải nguyên",
        "Negative", "High", "")

    add("GT-FI-017", "Ràng buộc", "Giá thanh toán âm → BLOCK",
        "Giá TT = -120.160",
        "1. Nhập giá âm\n2. Lưu",
        "Hệ thống CHẶN",
        "Negative", "High", "")

    add("GT-FI-018", "Mã GD", "Mã GD FI Bond format Fxxxxxxxxxx",
        "Tạo GD FI Bond",
        "1. Kiểm tra mã GD",
        "Mã bắt đầu bằng 'F': F0000000001",
        "Happy", "Medium", "")

    add("GT-FI-019", "Ngày tháng", "Kỳ hạn còn lại âm (CCTG đã đáo hạn)",
        "CCTG có Ngày đáo hạn < Ngày TT",
        "1. Ngày TT > Ngày đáo hạn\n2. Kiểm tra Kỳ hạn",
        "Kỳ hạn = số âm hoặc CẢNH BÁO",
        "Edge", "Medium", "")

    add("GT-FI-020", "Workflow", "CV sửa FI Bond sau TP duyệt → BLOCK",
        "GD ở Chờ K.NV duyệt",
        "1. CV cố sửa",
        "Hệ thống CHẶN",
        "Negative", "Critical", "")

    add("GT-FI-021", "Luồng duyệt", "CCTG qua đúng luồng GTCG",
        "GD CCTG",
        "1. CV → TP → GĐ → KTTC (2 cấp) → Hoàn thành",
        "CCTG = giống FI Bond. Không QLRR, không TTQT.",
        "Happy", "Critical", "")

    add("GT-FI-022", "Ràng buộc", "SL = 0 → BLOCK",
        "SL = 0",
        "1. SL = 0\n2. Lưu",
        "Hệ thống CHẶN",
        "Negative", "High", "")

    add("GT-FI-023", "Loại GD", "FI Bond Repo thành công",
        "Loại GD = Repo",
        "1. Chọn Repo\n2. Điền đầy đủ\n3. Lưu",
        "GD Repo thành công",
        "Happy", "Medium", "")

    add("GT-FI-024", "Ràng buộc", "Thiếu trường bắt buộc (Ngày đáo hạn)",
        "Bỏ trống Ngày đáo hạn",
        "1. Bỏ trống\n2. Lưu",
        "Hệ thống CHẶN: Ngày đáo hạn bắt buộc",
        "Negative", "High", "")

    add("GT-FI-025", "Tổng giá trị", "Tổng giá trị CCTG overflow check",
        "SL = 5.000.000; Giá TT = 1.000.000.000",
        "1. Nhập SL và Giá TT lớn\n2. Kiểm tra",
        "Tổng = 5.000.000.000.000.000 (5 triệu tỷ). Không overflow.",
        "Edge", "High", "")

    return tc


def mm_lien_nh():
    """M3-MM Liên ngân hàng test cases."""
    tc = []
    n = [0]
    def add(code, group, desc, precond, steps, expected, typ, sev, note=""):
        n[0] += 1
        tc.append([n[0], code, group, desc, precond, steps, expected, typ, sev, note])

    add("MM-LN-001", "Tạo GD", "Tạo GD Gửi tiền VND Liên NH thành công",
        "CV K.NV; Đối tác MSB; Loại tiền = VND",
        "1. Chiều = Gửi tiền\n2. Gốc = 100,000,000,000 VND\n3. LS = 4.50%/năm\n4. Day count = Actual/365\n5. Kỳ hạn = 90 ngày\n6. Lưu",
        "GD thành công.\nLãi = 100,000,000,000 × 4.50/100 × 90/365 = 1,109,589,041 VND (không thập phân)\nTổng đáo hạn = 101,109,589,041 VND",
        "Happy", "Critical", "VND: không thập phân")

    add("MM-LN-002", "Công thức", "Lãi USD với Actual/360",
        "USD; Gốc = 10,000,000.00; LS = 5.25%/năm; Day count = Actual/360; Kỳ hạn = 30 ngày",
        "1. Nhập đầy đủ\n2. Kiểm tra Lãi",
        "Lãi = 10,000,000 × 5.25/100 × 30/360 = 43,750.00 USD\nTổng = 10,043,750.00 USD (2 decimal)",
        "Happy", "Critical", "USD: 2 decimal; Day count = 360")

    add("MM-LN-003", "Công thức", "Lãi VND với Actual/360 (khác base)",
        "VND; Gốc = 500,000,000,000; LS = 3.80%/năm; Actual/360; 60 ngày",
        "1. Kiểm tra",
        "Lãi = 500B × 3.80/100 × 60/360 = 3,166,666,667 VND\n(làm tròn nguyên)",
        "Happy", "Critical", "Day count 360 vs 365 cho kết quả khác")

    add("MM-LN-004", "Công thức", "Lãi với Actual/Actual (năm nhuận 366 ngày)",
        "VND; 2028 là năm nhuận; Kỳ hạn nằm trong năm 2028",
        "1. Ngày hiệu lực = 01/01/2028; Kỳ hạn = 90\n2. Kiểm tra Day count base",
        "Day count base = 366 (năm nhuận).\nLãi tính với /366 thay vì /365",
        "Edge", "High", "Actual/Actual dùng số ngày thực tế trong năm")

    add("MM-LN-005", "Công thức", "Công thức ĐÚNG: Gốc × LS/100 × Kỳ hạn/N (không phải v1)",
        "Gốc = 100B; LS = 4.50%; Kỳ hạn = 90; Actual/365",
        "1. Kiểm tra công thức",
        "ĐÚNG: 100B × 4.50/100 × 90/365 = 1,109,589,041 VND\nSAI (v1): 100B × [(1+4.50/100) × 90]/365 = ??? (khác hoàn toàn)\nNếu kết quả khác → BUG công thức v1",
        "Happy", "Critical", "⚠️ BRD v1 có lỗi dấu ngoặc. Kiểm tra dùng đúng công thức v3")

    add("MM-LN-006", "Công thức", "Kỳ hạn = 1 ngày (overnight)",
        "Gốc = 200B VND; LS = 4.50%; Actual/365; Kỳ hạn = 1",
        "1. Kỳ hạn = 1 ngày\n2. Kiểm tra",
        "Lãi = 200B × 4.50/100 × 1/365 = 24,657,534 VND\nNgày đáo hạn = Ngày hiệu lực + 1 ngày",
        "Edge", "High", "Overnight")

    add("MM-LN-007", "Ngày tháng", "Ngày đáo hạn = Ngày hiệu lực + Kỳ hạn",
        "Ngày hiệu lực = 13/03/2026; Kỳ hạn = 90 ngày",
        "1. Kiểm tra Ngày đáo hạn tự động",
        "Ngày đáo hạn = 11/06/2026 (13/03 + 90 ngày)",
        "Happy", "High", "")

    add("MM-LN-008", "Ngày tháng", "Kỳ hạn = 0 → BLOCK",
        "Kỳ hạn = 0",
        "1. Nhập Kỳ hạn = 0\n2. Lưu",
        "Hệ thống CHẶN: Kỳ hạn phải > 0",
        "Negative", "High", "")

    add("MM-LN-009", "Ngày tháng", "Kỳ hạn âm → BLOCK",
        "Kỳ hạn = -30",
        "1. Nhập Kỳ hạn = -30",
        "Hệ thống CHẶN",
        "Negative", "High", "")

    add("MM-LN-010", "Định dạng", "VND không thập phân, USD 2 thập phân",
        "VND: Lãi = 1,109,589,041.23... ; USD: Lãi = 43,750.005...",
        "1. Kiểm tra hiển thị",
        "VND → 1,109,589,041 VND (làm tròn, không .xx)\nUSD → 43,750.01 USD (2 decimal)",
        "Edge", "Critical", "BRD: VND không thập phân, USD 2 decimal")

    add("MM-LN-011", "Luồng duyệt", "MM Liên NH CÓ qua QLRR (khác FX, GTCG)",
        "GD MM Liên NH",
        "1. CV → TP → GĐ → CV QLRR (cấp 1) → TPB QLRR (cấp 2) → KTTC (2 cấp) → TTQT (nếu có) → Hoàn thành",
        "MM Liên NH qua QLRR. Trạng thái 'Chờ QLRR' xuất hiện sau GĐ duyệt.",
        "Happy", "Critical", "CHỈ MM Liên NH qua QLRR")

    add("MM-LN-012", "Ràng buộc", "Gốc = 0 → BLOCK",
        "Gốc = 0",
        "1. Nhập Gốc = 0\n2. Lưu",
        "Hệ thống CHẶN: Số tiền gốc phải > 0",
        "Negative", "High", "")

    add("MM-LN-013", "Ràng buộc", "Gốc âm → BLOCK",
        "Gốc = -100,000,000,000",
        "1. Nhập gốc âm",
        "Hệ thống CHẶN",
        "Negative", "High", "")

    add("MM-LN-014", "Ràng buộc", "Lãi suất = 0%",
        "LS = 0.00%",
        "1. Nhập LS = 0\n2. Lưu",
        "GD tạo thành công. Lãi = 0. Tổng đáo hạn = Gốc.",
        "Edge", "Medium", "Cho phép LS = 0")

    add("MM-LN-015", "Ràng buộc", "Lãi suất âm → BLOCK hoặc cho phép?",
        "LS = -1.00%",
        "1. Nhập LS âm",
        "Hệ thống CHẶN: Lãi suất phải ≥ 0",
        "Negative", "Medium", "Tùy nghiệp vụ")

    add("MM-LN-016", "TSBĐ", "TSBĐ = Có → hiển thị 2 trường bổ sung",
        "TSBĐ = Có",
        "1. Chọn TSBĐ = Có\n2. Kiểm tra",
        "Hiển thị: Loại TSBĐ (dropdown VND/USD) + Giá trị TSBĐ (text)",
        "Happy", "High", "")

    add("MM-LN-017", "TSBĐ", "TSBĐ = Không → ẩn 2 trường bổ sung",
        "TSBĐ = Không",
        "1. Chọn TSBĐ = Không\n2. Kiểm tra",
        "2 trường Loại TSBĐ và Giá trị TSBĐ KHÔNG hiển thị",
        "Happy", "Medium", "")

    add("MM-LN-018", "Trường tự động", "Swift code tự động theo Đối tác",
        "Đối tác = MSB có SWIFT = MCOBVNVX",
        "1. Chọn Đối tác = MSB\n2. Kiểm tra Swift code",
        "Swift code = MCOBVNVX (tự động)",
        "Happy", "Medium", "")

    add("MM-LN-019", "Luồng duyệt", "CV QLRR từ chối hạn mức → Hủy giao dịch",
        "GD ở Chờ QLRR",
        "1. CV QLRR nhấn Không đồng ý\n2. Popup → lý do → Xác nhận",
        "GD = Hủy giao dịch. Notify CV K.NV.",
        "Happy", "Critical", "")

    add("MM-LN-020", "Luồng duyệt", "TPB QLRR từ chối hạn mức → Hủy giao dịch",
        "GD ở Chờ QLRR (CV QLRR đã duyệt cấp 1)",
        "1. TPB QLRR từ chối\n2. Popup → lý do → Xác nhận",
        "GD = Hủy giao dịch.",
        "Happy", "Critical", "")

    add("MM-LN-021", "Chỉ dẫn TT", "Chỉ dẫn TT đối tác tự động theo Đối tác + Loại tiền",
        "Đối tác MSB, Loại tiền VND",
        "1. Chọn Đối tác = MSB, Loại tiền = VND\n2. Kiểm tra Chỉ dẫn TT",
        "Chỉ dẫn TT tự động điền. Cho phép chỉnh sửa tay.",
        "Happy", "Medium", "")

    add("MM-LN-022", "Chiều GD", "Tạo GD 'Nhận tiền gửi' thành công",
        "Chiều = Nhận tiền gửi",
        "1. Chọn Chiều = Nhận tiền gửi\n2. Điền đầy đủ\n3. Lưu",
        "GD thành công. Chiều = Nhận tiền gửi.",
        "Happy", "High", "")

    add("MM-LN-023", "Chiều GD", "Tạo GD 'Cho vay' thành công",
        "Chiều = Cho vay",
        "1. Chọn Cho vay\n2. Lưu",
        "GD thành công",
        "Happy", "High", "")

    add("MM-LN-024", "Chiều GD", "Tạo GD 'Vay' thành công",
        "Chiều = Vay",
        "1. Chọn Vay\n2. Lưu",
        "GD thành công",
        "Happy", "High", "")

    add("MM-LN-025", "Công thức", "Division by zero: Day count = 0 (nếu có bug)",
        "Giả sử Day count base = 0 do lỗi",
        "1. Kiểm tra xử lý chia cho 0 trong công thức lãi",
        "Hệ thống KHÔNG crash. Hiển thị lỗi hoặc dùng default base.",
        "Edge", "High", "")

    add("MM-LN-026", "Workflow", "CV sửa MM sau TP duyệt → BLOCK",
        "GD ở Chờ K.NV duyệt",
        "1. CV cố sửa",
        "CHẶN",
        "Negative", "Critical", "")

    add("MM-LN-027", "Số lớn", "Gốc = 1 nghìn tỷ VND, kiểm tra overflow lãi",
        "Gốc = 1,000,000,000,000 VND; LS = 10%; 365 ngày; Actual/365",
        "1. Kiểm tra Lãi",
        "Lãi = 1,000,000,000,000 × 10/100 × 365/365 = 100,000,000,000 VND. Không overflow.",
        "Edge", "High", "")

    add("MM-LN-028", "Đáo hạn", "Hiển thị GD đến hạn trên danh sách (theo ngày đáo hạn)",
        "GD MM có Ngày đáo hạn = TODAY",
        "1. Kiểm tra danh sách MM\n2. Lọc theo Ngày đáo hạn = TODAY",
        "GD hiển thị trong danh sách đáo hạn hôm nay",
        "Happy", "High", "Chỉ hiển thị, không tự tạo GD mới")

    add("MM-LN-029", "TTQT", "MM Gửi tiền có TTQT = Có → qua TTQT",
        "GD Gửi tiền, TTQT = Có",
        "1. Hoàn thành luồng duyệt\n2. Kiểm tra TTQT",
        "GD xuất hiện trên TTQT khi Ngày hiệu lực = TODAY",
        "Happy", "High", "")

    add("MM-LN-030", "Qua năm", "GD MM qua năm (hiệu lực 2026, đáo hạn 2027)",
        "Ngày hiệu lực = 15/12/2026; Kỳ hạn = 90 ngày",
        "1. Kiểm tra Ngày đáo hạn",
        "Ngày đáo hạn = 15/03/2027. Tính đúng qua năm.",
        "Edge", "Medium", "")

    return tc


def mm_omo():
    """M3-MM OMO test cases."""
    tc = []
    n = [0]
    def add(code, group, desc, precond, steps, expected, typ, sev, note=""):
        n[0] += 1
        tc.append([n[0], code, group, desc, precond, steps, expected, typ, sev, note])

    add("MM-OMO-001", "Tạo GD", "Tạo GD OMO thành công",
        "CV K.NV; Đối tác tự động = 'Sở giao dịch NHNN'",
        "1. Phiên GD = Phiên 1\n2. Chọn Mã TP\n3. LS trúng thầu = 3.50%\n4. Kỳ hạn = 14 ngày\n5. Ngày TT 1 & TT 2\n6. Hair cut = 5%\n7. Lưu",
        "GD tạo thành công. Đối tác = Sở giao dịch NHNN (cố định).",
        "Happy", "Critical", "")

    add("MM-OMO-002", "Đối tác", "Đối tác OMO cố định = Sở giao dịch NHNN",
        "Tạo GD OMO",
        "1. Kiểm tra trường Đối tác",
        "Đối tác tự động = 'Sở giao dịch NHNN'. Không cho sửa.",
        "Happy", "High", "")

    add("MM-OMO-003", "Trường tự động", "TCPH, LS coupon, Ngày đáo hạn TP tự động từ danh mục",
        "Mã TP = TD2135068",
        "1. Chọn Mã TP\n2. Kiểm tra trường tự động",
        "TCPH, LS coupon, Ngày đáo hạn tự động điền từ danh mục",
        "Happy", "High", "")

    add("MM-OMO-004", "Luồng duyệt", "OMO KHÔNG qua QLRR, KHÔNG qua TTQT",
        "GD OMO",
        "1. CV → TP → GĐ → KTTC (2 cấp) → Hoàn thành",
        "KHÔNG có Chờ QLRR. KHÔNG có Chờ TTQT.",
        "Happy", "Critical", "BRD v3: OMO không qua QLRR, TTQT")

    add("MM-OMO-005", "Hạn mức", "OMO KHÔNG chiếm hạn mức",
        "GD OMO đã hoàn thành",
        "1. Kiểm tra Module 4 Hạn mức",
        "GD OMO KHÔNG xuất hiện trong hạn mức đã sử dụng",
        "Happy", "High", "BRD v3: OMO không chiếm hạn mức")

    add("MM-OMO-006", "Ngày tháng", "Ngày TT 2 phải sau Ngày TT 1",
        "Ngày TT 1 = 13/03/2026; Ngày TT 2 = 27/03/2026",
        "1. Nhập đúng thứ tự\n2. Lưu",
        "GD thành công",
        "Happy", "High", "")

    add("MM-OMO-007", "Ngày tháng", "Ngày TT 2 trước Ngày TT 1 → BLOCK",
        "Ngày TT 1 = 27/03/2026; Ngày TT 2 = 13/03/2026",
        "1. Nhập ngày TT 2 trước TT 1\n2. Lưu",
        "Hệ thống CHẶN: Ngày TT 2 phải sau Ngày TT 1",
        "Negative", "High", "")

    add("MM-OMO-008", "Ràng buộc", "Kỳ hạn = 0 → BLOCK",
        "Kỳ hạn = 0",
        "1. Nhập Kỳ hạn = 0",
        "CHẶN",
        "Negative", "High", "")

    add("MM-OMO-009", "Ràng buộc", "Hair cut = 0% (cho phép?)",
        "Hair cut = 0%",
        "1. Nhập Hair cut = 0\n2. Lưu",
        "Tùy nghiệp vụ: Cho phép hoặc cảnh báo",
        "Edge", "Medium", "")

    add("MM-OMO-010", "Ràng buộc", "Hair cut âm → BLOCK",
        "Hair cut = -5%",
        "1. Nhập Hair cut âm",
        "CHẶN",
        "Negative", "Medium", "")

    add("MM-OMO-011", "Ràng buộc", "LS trúng thầu = 0",
        "LS = 0",
        "1. Nhập LS = 0\n2. Lưu",
        "Cho phép (tùy nghiệp vụ) hoặc CẢNH BÁO",
        "Edge", "Medium", "")

    add("MM-OMO-012", "Ràng buộc", "LS trúng thầu âm → BLOCK",
        "LS = -3.50%",
        "1. Nhập LS âm",
        "CHẶN",
        "Negative", "Medium", "")

    add("MM-OMO-013", "Luồng duyệt", "TP trả lại → Open → CV sửa",
        "GD OMO vừa submit",
        "1. TP trả lại\n2. GD = Open\n3. CV sửa",
        "GD quay về Open, CV chỉnh sửa được",
        "Happy", "High", "")

    add("MM-OMO-014", "Luồng duyệt", "GĐ từ chối OMO → Popup → Từ chối",
        "GD Chờ K.NV duyệt",
        "1. GĐ từ chối\n2. Popup xác nhận",
        "GD = Từ chối. Notify CV.",
        "Happy", "Critical", "")

    add("MM-OMO-015", "Luồng duyệt", "CV KTTC cấp 1 từ chối OMO → Hủy giao dịch",
        "GD Chờ hạch toán",
        "1. CV KTTC từ chối\n2. Popup → Xác nhận",
        "GD = Hủy giao dịch.",
        "Happy", "Critical", "")

    add("MM-OMO-016", "Luồng duyệt", "LĐ KTTC cấp 2 từ chối OMO → Hủy giao dịch",
        "GD Chờ LĐ KTTC duyệt",
        "1. LĐ KTTC từ chối",
        "GD = Hủy giao dịch.",
        "Happy", "Critical", "")

    add("MM-OMO-017", "Workflow", "CV sửa OMO sau TP duyệt → BLOCK",
        "GD ở Chờ K.NV duyệt",
        "1. CV cố sửa",
        "CHẶN",
        "Negative", "Critical", "")

    add("MM-OMO-018", "Ràng buộc", "Thiếu Phiên giao dịch → BLOCK",
        "Bỏ trống Phiên GD",
        "1. Lưu không có Phiên GD",
        "CHẶN: Phiên GD bắt buộc",
        "Negative", "High", "")

    add("MM-OMO-019", "Trường tìm kiếm", "Dropdown Mã TP có tìm kiếm",
        "Danh mục TP",
        "1. Nhập 'TD' vào Mã TP",
        "Gợi ý danh sách TP bắt đầu 'TD'",
        "Happy", "Medium", "")

    add("MM-OMO-020", "File", "Upload file cho OMO",
        "File ticket",
        "1. Upload file\n2. Lưu",
        "File lưu thành công",
        "Happy", "Low", "")

    add("MM-OMO-021", "Ràng buộc", "Thiếu Mã TP → BLOCK",
        "Bỏ trống Mã TP",
        "1. Lưu",
        "CHẶN",
        "Negative", "High", "")

    add("MM-OMO-022", "Ngày tháng", "Ngày GD mặc định TODAY",
        "Mở form tạo OMO",
        "1. Kiểm tra Ngày GD",
        "Ngày GD = TODAY (mặc định)",
        "Happy", "Low", "")

    add("MM-OMO-023", "Trạng thái", "OMO không có trạng thái Chờ QLRR",
        "GD OMO đi qua tất cả bước",
        "1. Kiểm tra danh sách trạng thái",
        "Không tồn tại trạng thái 'Chờ QLRR' cho OMO",
        "Negative", "Critical", "Nếu có → BUG")

    add("MM-OMO-024", "Trạng thái", "OMO không có trạng thái Chờ TTQT",
        "GD OMO",
        "1. Kiểm tra",
        "Không tồn tại 'Chờ TTQT' cho OMO",
        "Negative", "Critical", "")

    add("MM-OMO-025", "Ngày tháng", "Ngày TT 1 = Ngày TT 2 → BLOCK hoặc cảnh báo",
        "TT 1 = TT 2 = 13/03/2026",
        "1. Nhập TT 1 = TT 2\n2. Lưu",
        "Hệ thống CHẶN hoặc CẢNH BÁO (kỳ hạn = 0 ngày)",
        "Edge", "Medium", "")

    return tc


def mm_repo_kbnn():
    """M3-MM Repo KBNN test cases."""
    tc = []
    n = [0]
    def add(code, group, desc, precond, steps, expected, typ, sev, note=""):
        n[0] += 1
        tc.append([n[0], code, group, desc, precond, steps, expected, typ, sev, note])

    add("MM-RK-001", "Tạo GD", "Tạo GD Repo KBNN thành công",
        "CV K.NV; Đối tác = Kho bạc Nhà nước",
        "1. Phiên GD = Phiên 1\n2. Đối tác = Kho bạc Nhà nước\n3. Chọn Mã TP\n4. LS trúng thầu = 3.50%\n5. Kỳ hạn = 14\n6. Ngày TT 1 & TT 2\n7. Hair cut = 5%\n8. Lưu",
        "GD thành công",
        "Happy", "Critical", "")

    add("MM-RK-002", "Luồng duyệt", "Repo KBNN KHÔNG qua QLRR, KHÔNG qua TTQT",
        "GD Repo KBNN",
        "1. CV → TP → GĐ → KTTC (2 cấp) → Hoàn thành",
        "KHÔNG Chờ QLRR. KHÔNG Chờ TTQT.",
        "Happy", "Critical", "Giống OMO")

    add("MM-RK-003", "Hạn mức", "Repo KBNN KHÔNG chiếm hạn mức",
        "GD Repo KBNN hoàn thành",
        "1. Kiểm tra Module 4",
        "GD Repo KBNN KHÔNG xuất hiện trong hạn mức",
        "Happy", "High", "")

    add("MM-RK-004", "Đối tác", "Đối tác Repo KBNN = dropdown (khác OMO cố định)",
        "Tạo GD Repo KBNN",
        "1. Kiểm tra trường Đối tác",
        "Đối tác = Dropdown theo danh sách (KHÔNG cố định như OMO)",
        "Happy", "High", "OMO: cố định NHNN; Repo KBNN: dropdown")

    add("MM-RK-005", "Ngày tháng", "Ngày TT 2 trước TT 1 → BLOCK",
        "TT 1 = 27/03; TT 2 = 13/03",
        "1. Nhập sai thứ tự\n2. Lưu",
        "CHẶN",
        "Negative", "High", "")

    add("MM-RK-006", "Ràng buộc", "Kỳ hạn = 0 → BLOCK",
        "Kỳ hạn = 0",
        "1. Nhập 0\n2. Lưu",
        "CHẶN",
        "Negative", "High", "")

    add("MM-RK-007", "Ràng buộc", "Thiếu trường bắt buộc (Mã TP)",
        "Bỏ trống Mã TP",
        "1. Lưu",
        "CHẶN",
        "Negative", "High", "")

    add("MM-RK-008", "Ràng buộc", "LS trúng thầu âm → BLOCK",
        "LS = -2%",
        "1. Nhập LS âm",
        "CHẶN",
        "Negative", "Medium", "")

    add("MM-RK-009", "Ràng buộc", "Hair cut âm → BLOCK",
        "Hair cut = -5%",
        "1. Nhập HC âm",
        "CHẶN",
        "Negative", "Medium", "")

    add("MM-RK-010", "Luồng duyệt", "TP trả lại → Open",
        "GD vừa submit",
        "1. TP trả lại",
        "GD = Open. CV sửa.",
        "Happy", "High", "")

    add("MM-RK-011", "Luồng duyệt", "GĐ từ chối → Popup → Từ chối",
        "GD Chờ K.NV duyệt",
        "1. GĐ từ chối\n2. Popup",
        "GD = Từ chối. Notify.",
        "Happy", "Critical", "")

    add("MM-RK-012", "Luồng duyệt", "CV KTTC từ chối → Hủy giao dịch",
        "GD Chờ hạch toán",
        "1. CV KTTC từ chối",
        "GD = Hủy giao dịch",
        "Happy", "Critical", "")

    add("MM-RK-013", "Luồng duyệt", "LĐ KTTC từ chối → Hủy giao dịch",
        "GD Chờ LĐ KTTC duyệt",
        "1. LĐ KTTC từ chối",
        "GD = Hủy giao dịch",
        "Happy", "Critical", "")

    add("MM-RK-014", "Workflow", "CV sửa sau TP duyệt → BLOCK",
        "GD Chờ K.NV duyệt",
        "1. CV cố sửa",
        "CHẶN",
        "Negative", "Critical", "")

    add("MM-RK-015", "Trạng thái", "Repo KBNN không có Chờ QLRR",
        "Theo dõi trạng thái",
        "1. Kiểm tra",
        "Không tồn tại Chờ QLRR",
        "Negative", "Critical", "")

    add("MM-RK-016", "Trạng thái", "Repo KBNN không có Chờ TTQT",
        "Theo dõi trạng thái",
        "1. Kiểm tra",
        "Không tồn tại Chờ TTQT",
        "Negative", "Critical", "")

    add("MM-RK-017", "Trường tự động", "TCPH, LS coupon tự động từ danh mục",
        "Chọn Mã TP",
        "1. Chọn TP từ dropdown\n2. Kiểm tra trường tự động",
        "Tự động điền TCPH, LS coupon, Ngày đáo hạn TP",
        "Happy", "High", "")

    add("MM-RK-018", "File", "Upload file cho Repo KBNN",
        "File hồ sơ",
        "1. Upload\n2. Lưu",
        "Lưu thành công",
        "Happy", "Low", "")

    add("MM-RK-019", "Ngày tháng", "Năm nhuận: Ngày TT 2 = 29/02/2028",
        "Năm 2028 nhuận",
        "1. Nhập TT 2 = 29/02/2028",
        "Hợp lệ. GD thành công.",
        "Edge", "Medium", "")

    add("MM-RK-020", "Ràng buộc", "Thiếu Phiên giao dịch",
        "Bỏ trống Phiên GD",
        "1. Lưu",
        "CHẶN. Thông báo bắt buộc nhập Phiên GD.",
        "Negative", "Medium", "")

    return tc


def han_muc():
    """Module 4: Hạn mức liên ngân hàng."""
    rows = []
    stt = [0]
    def add(ma, nhom, mota, dkdv, buoc, ketqua, loai, mucdo, ghichu):
        stt[0] += 1
        rows.append([stt[0], ma, nhom, mota, dkdv, buoc, ketqua, loai, mucdo, ghichu])

    # Happy cases
    add("HM-001", "Có TSBĐ", "Duyệt hạn mức có TSBĐ — luồng chính",
        "GD MM Liên NH, TSBĐ = Có, hạn mức chưa hết",
        "1. CV QLRR kiểm tra\n2. CV QLRR đồng ý\n3. TPB QLRR đồng ý",
        "GD chuyển Chờ hạch toán. Snapshot hạn mức lưu audit.",
        "Happy", "Critical", "")

    add("HM-002", "Không TSBĐ", "Duyệt hạn mức không TSBĐ — luồng chính",
        "GD MM Liên NH, TSBĐ = Không, hạn mức chưa hết",
        "1. CV QLRR kiểm tra\n2. CV QLRR đồng ý\n3. TPB QLRR đồng ý",
        "GD chuyển Chờ hạch toán.",
        "Happy", "Critical", "")

    add("HM-003", "Từ chối", "CV QLRR từ chối hạn mức",
        "GD Chờ QLRR",
        "1. CV QLRR từ chối\n2. Nhập lý do\n3. Popup xác nhận",
        "GD = Hủy giao dịch. Notify CV K.NV.",
        "Happy", "Critical", "")

    add("HM-004", "Từ chối", "TPB QLRR từ chối hạn mức",
        "CV QLRR đã duyệt",
        "1. TPB QLRR từ chối\n2. Popup xác nhận",
        "GD = Hủy giao dịch. Notify CV K.NV.",
        "Happy", "Critical", "")

    # Tính toán
    add("HM-005", "Tính toán", "Tỷ giá quy đổi = (Mua CK + Bán CK) / 2",
        "Tỷ giá mua CK = 25,900; Bán CK = 26,100",
        "1. Kiểm tra tỷ giá quy đổi hiển thị",
        "= (25,900 + 26,100) / 2 = 26,000",
        "Happy", "Critical", "Công thức BRD mục 3.4.3")

    add("HM-006", "Tính toán", "Hạn mức đã SD bao gồm cả FX + MM (v3)",
        "Đối tác có GD MM + FX chiếm hạn mức",
        "1. Kiểm tra cột Hạn mức đã sử dụng",
        "Tổng = GD MM chưa đáo hạn + GD FX chiếm hạn mức",
        "Happy", "Critical", "Thay đổi v3: FX chiếm hạn mức")

    add("HM-007", "Tính toán", "Giá trị còn lại = Hạn mức cấp − GD cần duyệt",
        "Hạn mức cấp = 500 tỷ; GD cần duyệt = 200 tỷ",
        "1. Kiểm tra trường Giá trị còn lại",
        "= 500 − 200 = 300 tỷ VND",
        "Happy", "High", "")

    add("HM-008", "Tính toán", "Hạn mức 'Không giới hạn' → kết quả = 'Không giới hạn'",
        "Hạn mức cấp (có TSBĐ) = Không giới hạn",
        "1. Kiểm tra Giá trị còn lại",
        "Hiển thị 'Không giới hạn' (KHÔNG tính toán số học)",
        "Edge", "Critical", "Lỗi thường gặp: tính toán với Infinity")

    add("HM-009", "Tính toán", "Hạn mức không TSBĐ bao gồm Giá thanh toán FI Bond",
        "Đối tác có GD FI Bond đã hạch toán, chưa đáo hạn, TCPH trùng Tên ĐT",
        "1. Kiểm tra Hạn mức đã SD không TSBĐ",
        "Bao gồm Giá thanh toán FI Bond (KHÔNG phải Số tiền gốc)",
        "Edge", "Critical", "Dùng Giá thanh toán, không phải Số tiền gốc")

    add("HM-010", "Tính toán", "Hạn mức đầu ngày: chỉ GD đáo hạn SAU ngày GD",
        "GD MM đáo hạn đúng ngày GD (hôm nay)",
        "1. Kiểm tra cột (6) (7) bảng tổng hợp",
        "GD đáo hạn hôm nay KHÔNG tính vào đã SD đầu ngày",
        "Edge", "Critical", "BRD: đáo hạn SAU ngày GD")

    add("HM-011", "Tính toán", "Quy đổi USD → VND khi tính hạn mức tổng",
        "GD MM: 10M USD + GD MM: 100 tỷ VND",
        "1. Kiểm tra tổng quy đổi VND",
        "= 10M × Tỷ giá QĐ + 100 tỷ VND",
        "Happy", "High", "")

    # Edge cases
    add("HM-012", "Ràng buộc", "GD vượt hạn mức còn lại (không TSBĐ)",
        "Hạn mức còn lại = 100 tỷ; GD cần duyệt = 150 tỷ",
        "1. Kiểm tra giá trị còn lại",
        "Giá trị còn lại = −50 tỷ (hiển thị âm hoặc cảnh báo)",
        "Edge", "Critical", "Hệ thống nên cảnh báo nhưng QLRR quyết định")

    add("HM-013", "Tính toán", "Chia đôi tỷ giá lẻ: (25,901 + 26,099) / 2",
        "Tỷ giá mua CK = 25,901; Bán CK = 26,099",
        "1. Kiểm tra tỷ giá quy đổi",
        "= 26,000 (không sai do làm tròn)",
        "Edge", "Medium", "")

    add("HM-014", "Ngày tháng", "Tỷ giá lấy cuối ngày LV liền trước (thứ 2 lấy thứ 6)",
        "GD tạo thứ Hai",
        "1. Kiểm tra tỷ giá quy đổi",
        "Lấy tỷ giá cuối ngày Thứ Sáu tuần trước",
        "Edge", "High", "")

    add("HM-015", "Bảng tổng hợp", "Xuất Excel bảng tổng hợp hạn mức",
        "Bảng tổng hợp có dữ liệu",
        "1. Nhấn xuất Excel",
        "File Excel đúng format, đủ 11 cột, số liệu khớp",
        "Happy", "High", "")

    add("HM-016", "Tính toán", "Bảng tổng hợp: Cột (10) = (4) − (6) − (8)",
        "Dữ liệu cột (4), (6), (8) đã có",
        "1. Kiểm tra giá trị cột (10)",
        "Đúng công thức: Hạn mức cấp − SD đầu ngày − SD trong ngày",
        "Happy", "Critical", "")

    add("HM-017", "Tính toán", "Bảng tổng hợp: Cột (11) Có TSBĐ = Không giới hạn",
        "Hạn mức có TSBĐ = Không giới hạn",
        "1. Kiểm tra cột (11)",
        "Hiển thị 'Không giới hạn'",
        "Edge", "Critical", "")

    add("HM-018", "Phân quyền", "CV QLRR chỉ thấy GD MM Liên NH Chờ QLRR",
        "Đăng nhập CV QLRR",
        "1. Xem danh sách GD",
        "Chỉ thấy GD MM Liên NH trạng thái Chờ QLRR",
        "Negative", "Critical", "")

    add("HM-019", "Phân quyền", "TPB QLRR chỉ thấy GD MM Liên NH Chờ QLRR",
        "Đăng nhập TPB QLRR",
        "1. Xem danh sách GD",
        "Chỉ thấy GD MM Liên NH trạng thái Chờ QLRR",
        "Negative", "Critical", "")

    add("HM-020", "Ràng buộc", "QLRR không duyệt được GD GTCG",
        "GD GTCG ở bất kỳ trạng thái nào",
        "1. CV QLRR cố duyệt GD GTCG",
        "CHẶN. GTCG không qua QLRR.",
        "Negative", "Critical", "GTCG, OMO, Repo KBNN không qua QLRR")

    add("HM-021", "Snapshot", "Snapshot hạn mức khi QLRR duyệt",
        "QLRR duyệt hạn mức thành công",
        "1. Kiểm tra audit trail",
        "Lưu snapshot: hạn mức cấp, đã SD, còn lại tại thời điểm duyệt",
        "Happy", "High", "BRD mục 8.3")

    add("HM-022", "Tính toán", "Nhiều GD cùng đối tác cùng ngày → cộng dồn",
        "Đối tác MSB có 3 GD MM trong ngày",
        "1. Kiểm tra GD cần duyệt",
        "Tổng = GD1 + GD2 + GD3",
        "Happy", "High", "")

    add("HM-023", "Ràng buộc", "Chỉ GD Gửi tiền/Cho vay chiếm hạn mức",
        "GD MM Chiều = Nhận tiền gửi",
        "1. Kiểm tra hạn mức",
        "GD Nhận tiền gửi/Vay KHÔNG chiếm hạn mức",
        "Edge", "Critical", "")

    add("HM-024", "Dữ liệu", "CIF tự động theo Đối tác",
        "Chọn đối tác MSB",
        "1. Kiểm tra CIF",
        "CIF tự động hiển thị theo đối tác",
        "Happy", "Medium", "")

    add("HM-025", "Ràng buộc", "Nhập thủ công hạn mức lần đầu, sau đó tự động",
        "Đối tác mới chưa có hạn mức",
        "1. Nhập hạn mức thủ công\n2. Lần sau kiểm tra",
        "Lần đầu cho nhập. Lần sau tự động + quyền sửa.",
        "Happy", "Medium", "")

    return rows


def ttqt():
    """Module 5: TTQT."""
    rows = []
    stt = [0]
    def add(ma, nhom, mota, dkdv, buoc, ketqua, loai, mucdo, ghichu):
        stt[0] += 1
        rows.append([stt[0], ma, nhom, mota, dkdv, buoc, ketqua, loai, mucdo, ghichu])

    # FX
    add("TT-001", "FX Spot", "GD FX Spot có TTQT → hiển thị đúng ngày thực hiện",
        "FX Spot, TTQT = Có, Ngày TH = TODAY, TT = Chờ TTQT",
        "1. Đăng nhập BP.TTQT\n2. Xem danh sách",
        "GD hiển thị. Số tiền = Khối lượng GD. Ngày chuyển = Ngày TH.",
        "Happy", "Critical", "")

    add("TT-002", "FX Spot", "BP.TTQT duyệt → Hoàn thành",
        "GD FX Chờ TTQT",
        "1. BP.TTQT chọn 'Có'\n2. Xác nhận",
        "GD = Hoàn thành.",
        "Happy", "Critical", "")

    add("TT-003", "FX Spot", "BP.TTQT từ chối → Hủy + popup xác nhận",
        "GD FX Chờ TTQT",
        "1. BP.TTQT chọn 'Không'\n2. Nhập lý do\n3. Popup xác nhận\n4. Xác nhận",
        "GD = Hủy giao dịch. Notify CV K.NV.",
        "Happy", "Critical", "")

    add("TT-004", "FX Forward", "GD Forward TTQT chỉ hiển thị đúng Ngày thực hiện",
        "FX Forward, TTQT = Có, Ngày TH = 15/05/2026",
        "1. Vào TTQT ngày 14/05\n2. Vào TTQT ngày 15/05",
        "Ngày 14: KHÔNG hiển thị. Ngày 15: hiển thị.",
        "Edge", "Critical", "GD chỉ hiện đúng ngày")

    # Swap
    add("TT-005", "FX Swap", "Swap chân 1: Ticket suffix = 'A'",
        "Swap, TTQT chân 1 = Có, Ticket = 12345",
        "1. Kiểm tra danh sách TTQT",
        "Số Ticket = '12345A'",
        "Happy", "Critical", "BRD: chân 1 suffix A")

    add("TT-006", "FX Swap", "Swap chân 2: Ticket suffix = 'B', auto hiện đúng ngày",
        "Swap, TTQT chân 2 = Có, Ngày TH chân 2 = 15/06/2026",
        "1. Vào TTQT ngày 15/06",
        "GD tự động hiển thị. Ticket = '12345B'. Số tiền = Khối lượng GD.",
        "Happy", "Critical", "Chân 2 TỰ ĐỘNG xuất hiện")

    add("TT-007", "FX Swap", "Swap chân 2 KHÔNG hiện trước ngày thực hiện",
        "Swap chân 2, Ngày TH chân 2 = 15/06/2026",
        "1. Vào TTQT ngày 14/06",
        "Chân 2 CHƯA hiển thị.",
        "Edge", "Critical", "")

    # MM
    add("TT-008", "MM Gửi/Cho vay", "MM Gửi tiền: hiển thị TTQT ngày hiệu lực",
        "MM Liên NH, Chiều = Gửi tiền, TTQT = Có, Ngày HL = TODAY",
        "1. Xem TTQT",
        "Hiển thị. Số tiền = Số tiền gốc. Ngày chuyển = Ngày hiệu lực.",
        "Happy", "Critical", "")

    add("TT-009", "MM Nhận/Vay", "MM Nhận tiền gửi: TTQT ngày đáo hạn, số tiền = Gốc + Lãi",
        "MM Liên NH, Chiều = Nhận tiền gửi, TTQT = Có, Ngày ĐH = TODAY",
        "1. Xem TTQT",
        "Hiển thị. Số tiền = Số tiền tại ngày đáo hạn (Gốc + Lãi).",
        "Happy", "Critical", "KHÔNG phải Gốc. Phải là Gốc + Lãi")

    add("TT-010", "MM Cho vay", "MM Cho vay USD: format 2 số thập phân",
        "MM Cho vay, USD, 10,000,000.00 USD",
        "1. Kiểm tra Số tiền trên TTQT",
        "10,000,000.00 USD (2 decimal)",
        "Happy", "High", "")

    add("TT-011", "Ràng buộc", "GD GTCG KHÔNG hiển thị trên TTQT",
        "GD GTCG Hoàn thành",
        "1. Kiểm tra TTQT",
        "KHÔNG có GD GTCG. GTCG không qua TTQT.",
        "Negative", "Critical", "")

    add("TT-012", "Ràng buộc", "GD OMO KHÔNG hiển thị trên TTQT",
        "GD OMO Hoàn thành",
        "1. Kiểm tra TTQT",
        "KHÔNG có GD OMO.",
        "Negative", "Critical", "OMO không qua TTQT")

    add("TT-013", "Ràng buộc", "GD Repo KBNN KHÔNG hiển thị trên TTQT",
        "GD Repo KBNN Hoàn thành",
        "1. Kiểm tra TTQT",
        "KHÔNG có GD Repo KBNN.",
        "Negative", "Critical", "")

    add("TT-014", "Phân quyền", "BP.TTQT chỉ thấy GD Chờ TTQT đến hạn TODAY",
        "Đăng nhập BP.TTQT",
        "1. Xem danh sách",
        "Chỉ thấy GD có Chờ TTQT + Ngày = TODAY",
        "Negative", "Critical", "")

    add("TT-015", "Mapping", "BIC CODE tự động theo Tên đối tác",
        "GD FX có đối tác MSB",
        "1. Kiểm tra BIC CODE trên TTQT",
        "Tự động hiển thị SWIFT code MSB",
        "Happy", "High", "")

    add("TT-016", "Mapping", "Trích tiền từ tài khoản = TK HABIB KLB",
        "GD FX Chờ TTQT",
        "1. Kiểm tra trường 'Trích tiền từ tài khoản'",
        "= Tài khoản HABIB của KienlongBank",
        "Happy", "High", "")

    add("TT-017", "Workflow", "Lọc mặc định TTQT = TODAY",
        "Nhiều GD TTQT nhiều ngày",
        "1. Mở màn hình TTQT",
        "Mặc định lọc Ngày GD = TODAY",
        "Happy", "Medium", "")

    add("TT-018", "Ràng buộc", "Citad KHÔNG đi qua Treasury",
        "GD nội địa qua Citad",
        "1. Kiểm tra TTQT",
        "Citad tự động trên Core. KHÔNG hiển thị trên Treasury.",
        "Negative", "High", "BRD: Citad không qua Treasury")

    add("TT-019", "Workflow", "Hủy TTQT → BP.TTQT nhấn Không → popup",
        "GD Chờ TTQT",
        "1. Nhấn 'Không'\n2. Popup xác nhận\n3. Nhấn 'Hủy' trên popup",
        "Quay lại. GD giữ nguyên Chờ TTQT.",
        "Edge", "High", "Chống click nhầm")

    add("TT-020", "Workflow", "MM Vay: TTQT ngày đáo hạn (trả tiền)",
        "MM Chiều = Vay, TTQT = Có, Ngày ĐH = 15/06/2026",
        "1. Vào TTQT ngày 15/06",
        "Hiển thị. Số tiền = Gốc + Lãi. Ngày chuyển = Ngày ĐH.",
        "Happy", "Critical", "")

    add("TT-021", "Danh sách", "Cột Loại GD hiển thị 'FX' hoặc 'MM'",
        "Danh sách TTQT có cả FX và MM",
        "1. Kiểm tra cột Loại giao dịch",
        "FX → 'FX'; MM → 'MM'",
        "Happy", "Medium", "")

    add("TT-022", "Edge", "Swap 2 chân cùng ngày TH: cả 2 hiện trên TTQT",
        "Swap, Ngày TH chân 1 = Ngày TH chân 2 = TODAY",
        "1. Xem TTQT",
        "Hiển thị 2 dòng: Ticket 'A' và Ticket 'B'",
        "Edge", "High", "")

    add("TT-023", "Dữ liệu", "SSI đối tác mapping đúng từ Pay code",
        "FX Spot, Pay code ĐT = 'So 8901... IRVTUS3N'",
        "1. Kiểm tra SSI đối tác trên TTQT",
        "SSI = Pay code đối tác từ FX",
        "Happy", "High", "")

    add("TT-024", "Lọc", "Lọc theo Loại tiền, Đối tác, Trạng thái",
        "Danh sách TTQT có nhiều GD",
        "1. Lọc USD\n2. Lọc theo tên ĐT\n3. Lọc Hoàn thành",
        "Kết quả đúng filter",
        "Happy", "Medium", "")

    add("TT-025", "Edge", "GD FX TTQT = Không → KHÔNG hiển thị",
        "FX Spot, TTQT = Không, Hoàn thành",
        "1. Kiểm tra TTQT",
        "KHÔNG hiển thị. GD hoàn thành trực tiếp sau KTTC duyệt.",
        "Negative", "High", "")

    return rows


def luong_dac_biet():
    """Luồng đặc biệt: Recall, Hủy, Clone."""
    rows = []
    stt = [0]
    def add(ma, nhom, mota, dkdv, buoc, ketqua, loai, mucdo, ghichu):
        stt[0] += 1
        rows.append([stt[0], ma, nhom, mota, dkdv, buoc, ketqua, loai, mucdo, ghichu])

    # Recall
    add("LDB-001", "Recall", "CV Recall khi GD Chờ K.NV duyệt → Open",
        "GD Chờ K.NV duyệt (TP đã duyệt, GĐ chưa duyệt)",
        "1. CV nhấn Recall\n2. Nhập lý do\n3. Xác nhận",
        "GD quay về Open. CV có thể chỉnh sửa.",
        "Happy", "Critical", "")

    add("LDB-002", "Recall", "TP Recall khi GD Chờ K.NV duyệt → quay về TP",
        "GD Chờ K.NV duyệt (TP đã duyệt, GĐ chưa duyệt)",
        "1. TP nhấn Recall\n2. Nhập lý do\n3. Xác nhận",
        "GD quay về bước TP review lại (KHÔNG về Open).",
        "Happy", "Critical", "TP recall ≠ CV recall")

    add("LDB-003", "Recall", "Recall BẮT BUỘC nhập lý do",
        "GD Chờ K.NV duyệt",
        "1. CV nhấn Recall\n2. Bỏ trống lý do\n3. Xác nhận",
        "CHẶN. Bắt buộc nhập lý do recall.",
        "Negative", "High", "")

    add("LDB-004", "Recall", "Recall khi GD đã Hoàn thành → CHẶN",
        "GD Hoàn thành",
        "1. CV cố recall",
        "CHẶN. Recall chỉ khi Chờ K.NV duyệt.",
        "Negative", "Critical", "Recall ≠ Hủy")

    add("LDB-005", "Recall", "Recall khi GD Open → CHẶN",
        "GD Open (chưa submit)",
        "1. CV cố recall",
        "CHẶN. Không có nút Recall khi Open.",
        "Negative", "High", "")

    add("LDB-006", "Recall", "Recall khi GD Chờ hạch toán → CHẶN",
        "GD Chờ hạch toán (GĐ đã duyệt)",
        "1. CV cố recall",
        "CHẶN. Recall chỉ áp dụng khi Chờ K.NV duyệt.",
        "Negative", "Critical", "")

    add("LDB-007", "Recall", "Audit trail ghi nhận recall",
        "CV recall thành công",
        "1. Xem lịch sử GD",
        "Log: User, timestamp, lý do, trạng thái trước/sau.",
        "Happy", "High", "")

    # Hủy
    add("LDB-008", "Hủy", "Hủy GD Hoàn thành: CV → TP (cấp 1) → GĐ (cấp 2)",
        "GD Hoàn thành",
        "1. CV yêu cầu hủy + lý do\n2. TP duyệt hủy cấp 1\n3. GĐ/PGĐ Khối duyệt hủy cấp 2",
        "GD = Đã hủy. Email notify P.KTTC.",
        "Happy", "Critical", "Hủy cần 2 cấp duyệt (v3)")

    add("LDB-009", "Hủy", "Hủy GD Chờ TTQT",
        "GD Chờ TTQT",
        "1. CV yêu cầu hủy\n2. TP duyệt\n3. GĐ duyệt",
        "GD = Đã hủy. Email notify P.KTTC + BP.TTQT.",
        "Happy", "Critical", "GD có TTQT → email thêm BP.TTQT")

    add("LDB-010", "Hủy", "TP từ chối hủy → GD giữ nguyên",
        "GD Hoàn thành, CV yêu cầu hủy",
        "1. TP từ chối hủy cấp 1",
        "GD giữ nguyên Hoàn thành.",
        "Happy", "High", "")

    add("LDB-011", "Hủy", "GĐ từ chối hủy → GD giữ nguyên",
        "TP đã duyệt hủy cấp 1",
        "1. GĐ từ chối hủy cấp 2",
        "GD giữ nguyên Hoàn thành.",
        "Happy", "High", "")

    add("LDB-012", "Hủy", "GD Đã hủy vẫn hiển thị (mặc định ẩn, filter hiện)",
        "GD Đã hủy",
        "1. Xem danh sách GD (mặc định)\n2. Bật filter hiện GD hủy",
        "Mặc định: ẩn GD hủy. Bật filter: hiển thị với trạng thái 'Đã hủy'.",
        "Happy", "High", "KHÔNG xóa khỏi hệ thống")

    add("LDB-013", "Hủy", "Hủy GD Open → CHẶN",
        "GD Open",
        "1. CV cố hủy",
        "CHẶN. Hủy chỉ áp dụng cho GD Hoàn thành hoặc Chờ TTQT.",
        "Negative", "Critical", "")

    add("LDB-014", "Hủy", "Hủy GD Từ chối → CHẶN",
        "GD Từ chối",
        "1. CV cố hủy",
        "CHẶN. GD Từ chối dùng Clone, không phải Hủy.",
        "Negative", "High", "")

    add("LDB-015", "Hủy", "Email thông báo khi hủy GD không có TTQT",
        "GD Hoàn thành, TTQT = Không",
        "1. Hủy thành công",
        "Email CHỈ gửi P.KTTC. KHÔNG gửi BP.TTQT.",
        "Happy", "High", "")

    # Clone
    add("LDB-016", "Clone", "Clone GD bị Từ chối → GD mới Open",
        "GD Từ chối (GĐ từ chối)",
        "1. CV nhấn Clone",
        "GD mới tạo, copy dữ liệu cũ, trạng thái Open.",
        "Happy", "Critical", "")

    add("LDB-017", "Clone", "Clone GD bị Hủy giao dịch (KTTC từ chối)",
        "GD Hủy giao dịch (P.KTTC từ chối)",
        "1. CV nhấn Clone",
        "GD mới tạo, copy dữ liệu cũ, trạng thái Open.",
        "Happy", "Critical", "")

    add("LDB-018", "Clone", "Clone GD Hoàn thành → CHẶN",
        "GD Hoàn thành",
        "1. CV cố Clone",
        "CHẶN. Clone chỉ cho GD Từ chối hoặc Hủy GD (KTTC từ chối).",
        "Negative", "Critical", "")

    add("LDB-019", "Clone", "Clone GD Đã hủy (hủy sau hoàn thành) → CHẶN",
        "GD Đã hủy (đã hoàn thành rồi hủy)",
        "1. CV cố Clone",
        "CHẶN. Clone chỉ áp dụng cho Từ chối/Hủy GD (KTTC từ chối).",
        "Negative", "High", "Phân biệt 'Hủy giao dịch' vs 'Đã hủy'")

    add("LDB-020", "Clone", "GD clone có ID mới, dữ liệu copy đầy đủ",
        "GD Từ chối, có đầy đủ dữ liệu",
        "1. Clone\n2. Kiểm tra GD mới",
        "ID mới. Tất cả trường dữ liệu copy chính xác. Trạng thái = Open.",
        "Happy", "High", "")

    add("LDB-021", "Clone", "CV chỉnh sửa GD clone trước khi submit",
        "GD clone vừa tạo (Open)",
        "1. Sửa tỷ giá\n2. Submit",
        "Cho phép sửa. Submit bình thường.",
        "Happy", "Medium", "")

    # Popup
    add("LDB-022", "Popup", "Popup xác nhận khi GĐ từ chối",
        "GD Chờ K.NV duyệt",
        "1. GĐ nhấn Từ chối\n2. Popup hiện\n3. Nhập lý do\n4. Xác nhận",
        "Popup bắt buộc. Có nút 'Xác nhận từ chối' + 'Hủy'.",
        "Happy", "Critical", "Chống click nhầm")

    add("LDB-023", "Popup", "Popup xác nhận khi KTTC từ chối",
        "GD Chờ hạch toán",
        "1. CV KTTC nhấn Từ chối\n2. Popup hiện",
        "Popup bắt buộc nhập lý do.",
        "Happy", "Critical", "")

    add("LDB-024", "Popup", "Nhấn 'Hủy' trên popup → quay lại, GD không thay đổi",
        "GD Chờ K.NV duyệt, GĐ nhấn Từ chối",
        "1. Popup hiện\n2. Nhấn 'Hủy'",
        "Quay lại. GD giữ nguyên Chờ K.NV duyệt.",
        "Edge", "High", "")

    add("LDB-025", "Popup", "Popup bắt buộc lý do không để trống",
        "Popup từ chối hiện",
        "1. Bỏ trống lý do\n2. Nhấn Xác nhận",
        "CHẶN. Bắt buộc nhập lý do.",
        "Negative", "High", "")

    # Cross-module
    add("LDB-026", "Cross", "Recall áp dụng cho FX, GTCG, MM",
        "GD FX Chờ K.NV duyệt",
        "1. CV recall GD FX\n2. CV recall GD GTCG\n3. CV recall GD MM",
        "Cả 3 module đều hỗ trợ recall.",
        "Happy", "High", "Luồng đặc biệt áp dụng chung")

    add("LDB-027", "Cross", "Hủy áp dụng cho FX, GTCG, MM",
        "GD Hoàn thành các module",
        "1. Hủy GD FX\n2. Hủy GD GTCG\n3. Hủy GD MM",
        "Cả 3 module đều hỗ trợ hủy.",
        "Happy", "High", "")

    add("LDB-028", "Audit", "Tất cả thao tác đặc biệt đều ghi audit trail",
        "Recall / Hủy / Clone",
        "1. Thực hiện\n2. Kiểm tra lịch sử",
        "Audit trail đầy đủ: user, timestamp, hành động, lý do.",
        "Happy", "High", "")

    add("LDB-029", "Edge", "Recall rồi submit lại → luồng bình thường",
        "GD đã recall về Open",
        "1. CV sửa\n2. Submit lại\n3. TP duyệt\n4. GĐ duyệt",
        "Luồng chạy bình thường sau recall.",
        "Happy", "Medium", "")

    add("LDB-030", "Edge", "Clone GD FX Swap → copy đủ 2 chân",
        "GD FX Swap bị Từ chối",
        "1. Clone",
        "GD mới có đầy đủ dữ liệu chân 1 + chân 2.",
        "Edge", "High", "")

    return rows


def phan_quyen_audit():
    """Phân quyền & Audit trail."""
    rows = []
    stt = [0]
    def add(ma, nhom, mota, dkdv, buoc, ketqua, loai, mucdo, ghichu):
        stt[0] += 1
        rows.append([stt[0], ma, nhom, mota, dkdv, buoc, ketqua, loai, mucdo, ghichu])

    # Phân quyền theo role
    add("PQ-001", "CV K.NV", "CV tạo GD FX, GTCG, MM",
        "Đăng nhập CV K.NV",
        "1. Tạo GD FX\n2. Tạo GD GTCG\n3. Tạo GD MM",
        "Tạo thành công tất cả module.",
        "Happy", "Critical", "")

    add("PQ-002", "CV K.NV", "CV sửa GD khi Open",
        "GD Open, đăng nhập CV K.NV",
        "1. Sửa trường dữ liệu\n2. Lưu",
        "Lưu thành công.",
        "Happy", "Critical", "")

    add("PQ-003", "CV K.NV", "CV KHÔNG sửa được GD sau TP duyệt",
        "GD Chờ K.NV duyệt",
        "1. CV cố sửa trường dữ liệu",
        "CHẶN. Không có quyền chỉnh sửa sau TP duyệt.",
        "Negative", "Critical", "Quy tắc BRD: TP duyệt → lock")

    add("PQ-004", "TP K.NV", "TP duyệt cấp 1 + Recall",
        "GD Open đã submit",
        "1. TP xem GD\n2. Duyệt hoặc Trả lại",
        "TP có quyền Duyệt cấp 1 + Trả lại + Recall.",
        "Happy", "Critical", "")

    add("PQ-005", "GĐ TT", "GĐ TT KDV/QLV duyệt cấp 2",
        "GD Chờ K.NV duyệt",
        "1. GĐ TT duyệt",
        "GD chuyển trạng thái tiếp theo.",
        "Happy", "Critical", "")

    add("PQ-006", "GĐ/PGĐ Khối", "GĐ/PGĐ Khối duyệt cấp 2 + Duyệt hủy",
        "GD Chờ K.NV duyệt",
        "1. GĐ Khối duyệt cấp 2\n2. GĐ Khối duyệt hủy (GD Hoàn thành)",
        "Có quyền duyệt cấp 2 + duyệt hủy.",
        "Happy", "Critical", "")

    add("PQ-007", "CV QLRR", "CV QLRR CHỈ thấy GD MM Liên NH Chờ QLRR",
        "Đăng nhập CV QLRR",
        "1. Xem tất cả module",
        "Chỉ thấy GD MM Liên NH trạng thái Chờ QLRR. KHÔNG thấy FX, GTCG, OMO.",
        "Negative", "Critical", "")

    add("PQ-008", "CV KTTC", "CV KTTC CHỈ thấy GD từ bước hạch toán trở đi",
        "Đăng nhập CV P.KTTC",
        "1. Xem danh sách GD",
        "Chỉ thấy GD Chờ hạch toán + các trạng thái sau. KHÔNG thấy Open, Chờ K.NV duyệt.",
        "Negative", "Critical", "")

    add("PQ-009", "LĐ KTTC", "LĐ KTTC duyệt hạch toán cấp 2",
        "GD Chờ LĐ KTTC duyệt (CV KTTC đã duyệt cấp 1)",
        "1. LĐ KTTC duyệt",
        "GD chuyển Chờ TTQT hoặc Hoàn thành.",
        "Happy", "Critical", "KTTC 2 cấp (v3)")

    add("PQ-010", "BP.TTQT", "BP.TTQT CHỈ thấy GD Chờ TTQT đến hạn TODAY",
        "Đăng nhập BP.TTQT",
        "1. Xem danh sách",
        "Chỉ thấy GD Chờ TTQT ngày hôm nay.",
        "Negative", "Critical", "")

    add("PQ-011", "Admin", "Admin quản lý Master Data",
        "Đăng nhập Admin",
        "1. Thêm đối tác\n2. Import danh mục TP\n3. Cập nhật Pay code",
        "Thao tác thành công.",
        "Happy", "High", "")

    add("PQ-012", "Admin", "Admin xem tất cả GD nhưng KHÔNG duyệt",
        "Đăng nhập Admin",
        "1. Xem GD\n2. Cố duyệt",
        "Xem: OK. Duyệt: CHẶN.",
        "Negative", "High", "")

    # Self-approval
    add("PQ-013", "Self-approve", "Cùng người tạo và duyệt (CV = TP) → CHẶN",
        "User có role CV, GD do user tạo",
        "1. Đăng nhập role TP cùng user\n2. Cố duyệt GD mình tạo",
        "CHẶN. Không được tự duyệt GD của mình.",
        "Negative", "Critical", "Kiểm soát nội bộ")

    add("PQ-014", "Self-approve", "TP tạo GD rồi tự duyệt cấp 1 → CHẶN",
        "TP tạo GD (nếu TP có quyền tạo)",
        "1. TP tạo GD\n2. TP cố duyệt cấp 1",
        "CHẶN.",
        "Negative", "Critical", "")

    # Ủy quyền
    add("PQ-015", "Ủy quyền", "Người được ủy quyền dùng account riêng",
        "Người A ủy quyền Người B",
        "1. Người B đăng nhập account riêng\n2. Thao tác",
        "Account riêng, đăng nhập trực tiếp. KHÔNG có chức năng ủy quyền trên hệ thống.",
        "Happy", "Medium", "BRD: không cần ủy quyền trên HT")

    # Audit trail
    add("PQ-016", "Audit", "Tạo GD → ghi log audit",
        "CV tạo GD mới",
        "1. Tạo GD\n2. Xem lịch sử",
        "Log: User ID, Họ tên, Đơn vị, Timestamp, Sự kiện = Tạo GD.",
        "Happy", "Critical", "")

    add("PQ-017", "Audit", "Sửa GD → ghi log giá trị cũ + mới",
        "CV sửa tỷ giá từ 26,000 → 26,100",
        "1. Sửa\n2. Xem lịch sử",
        "Log: Giá trị cũ = 26,000; Giá trị mới = 26,100.",
        "Happy", "Critical", "")

    add("PQ-018", "Audit", "Từ chối → log lý do + trạng thái trước/sau",
        "GĐ từ chối GD",
        "1. Xem lịch sử",
        "Log: Trạng thái trước = Chờ K.NV duyệt, Trạng thái sau = Từ chối, Lý do = [text].",
        "Happy", "Critical", "")

    add("PQ-019", "Audit", "Log append-only: KHÔNG cho sửa/xóa",
        "Audit log đã tạo",
        "1. Cố sửa log\n2. Cố xóa log",
        "CHẶN. Audit trail chỉ append, không sửa/xóa.",
        "Negative", "Critical", "BRD mục 8.3")

    add("PQ-020", "Audit", "Snapshot hạn mức khi QLRR duyệt",
        "QLRR duyệt hạn mức thành công",
        "1. Xem audit trail",
        "Snapshot: hạn mức cấp, đã SD, còn lại tại thời điểm duyệt.",
        "Happy", "High", "")

    add("PQ-021", "Audit", "Mỗi GD có tab Lịch sử hiển thị timeline",
        "GD đã qua nhiều bước",
        "1. Mở GD\n2. Vào tab Lịch sử",
        "Timeline đầy đủ các hành động theo thứ tự thời gian.",
        "Happy", "High", "")

    add("PQ-022", "Audit", "Quyền xem lịch sử: TP K.NV trở lên + QLRR + KTTC",
        "Đăng nhập CV K.NV",
        "1. CV K.NV xem lịch sử",
        "CV K.NV: kiểm tra BRD — TP K.NV trở lên mới xem được.",
        "Edge", "Medium", "BRD mục 8.3")

    add("PQ-023", "Audit", "Lưu trữ vĩnh viễn, không purge",
        "Audit log cũ > 1 năm",
        "1. Kiểm tra log cũ",
        "Vẫn truy xuất được. Không có cơ chế purge.",
        "Happy", "Medium", "")

    add("PQ-024", "Audit", "14 loại sự kiện đều ghi log",
        "Thực hiện lần lượt 14 loại sự kiện",
        "1. Tạo, Sửa, Submit, Duyệt C1, Duyệt C2, QLRR, KTTC C1, KTTC C2, TTQT, Recall, Yêu cầu hủy, Duyệt hủy C1, Duyệt hủy C2, Clone",
        "Tất cả 14 sự kiện đều có log entry.",
        "Happy", "Critical", "BRD mục 8.1")

    add("PQ-025", "Notification", "Chuyển trạng thái → In-app notify người bước tiếp",
        "GD chuyển từ Open → Chờ K.NV duyệt",
        "1. Kiểm tra notification",
        "In-app notify GĐ TT/GĐ Khối.",
        "Happy", "High", "")

    add("PQ-026", "Notification", "GĐ từ chối → In-app notify CV K.NV",
        "GĐ từ chối GD",
        "1. Kiểm tra notification CV",
        "CV nhận in-app notification.",
        "Happy", "High", "")

    add("PQ-027", "Notification", "Hủy GD có TTQT → Email cả KTTC + TTQT",
        "Hủy GD đã hạch toán, TTQT = Có",
        "1. Hủy thành công\n2. Kiểm tra email",
        "Email gửi P.KTTC + BP.TTQT.",
        "Happy", "Critical", "v3: email thêm TTQT")

    add("PQ-028", "Notification", "Hủy GD không TTQT → Email CHỈ KTTC",
        "Hủy GD đã hạch toán, TTQT = Không",
        "1. Hủy thành công\n2. Kiểm tra email",
        "Email CHỈ gửi P.KTTC. KHÔNG gửi TTQT.",
        "Happy", "High", "")

    add("PQ-029", "Xem dữ liệu", "K.NV (CV/TP/GĐ) xem toàn bộ GD",
        "Đăng nhập CV K.NV",
        "1. Xem GD FX, GTCG, MM",
        "Xem được toàn bộ GD tất cả module.",
        "Happy", "High", "")

    add("PQ-030", "Xem dữ liệu", "QLRR KHÔNG xem được GD FX, GTCG",
        "Đăng nhập CV QLRR",
        "1. Cố truy cập FX, GTCG",
        "CHẶN. QLRR chỉ xem MM Liên NH cần duyệt hạn mức.",
        "Negative", "Critical", "")

    return rows


# ── MAIN ──
def main():
    wb = openpyxl.Workbook()

    # Sheet 1: Hướng dẫn
    create_guide_sheet(wb)

    # Sheet 2-12: Test cases
    sheets_config = [
        ("M1-FX Spot Forward", fx_spot_forward),
        ("M1-FX Swap", fx_swap),
        ("M2-GTCG Govi Bond", gtcg_govi),
        ("M2-GTCG FI Bond CCTG", gtcg_fi_cctg),
        ("M3-MM Liên NH", mm_lien_nh),
        ("M3-MM OMO", mm_omo),
        ("M3-MM Repo KBNN", mm_repo_kbnn),
        ("M4-Hạn mức", han_muc),
        ("M5-TTQT", ttqt),
        ("Luồng đặc biệt", luong_dac_biet),
        ("Phân quyền Audit", phan_quyen_audit),
    ]

    total = 0
    for sheet_name, data_func in sheets_config:
        ws = wb.create_sheet(title=sheet_name)
        data_rows = data_func()
        total += len(data_rows)

        # Write header
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # Write data
        for row_idx, row_data in enumerate(data_rows, 2):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = WRAP_ALIGN if col_idx > 2 else CENTER_ALIGN

        # Apply formatting
        style_sheet(ws, data_rows)

        # Column widths
        for col_idx, width in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header
        ws.freeze_panes = "A2"

    # Update total count on guide sheet
    guide = wb["Hướng dẫn"]
    for row in guide.iter_rows(min_row=1, max_row=30, min_col=2, max_col=3):
        for cell in row:
            if cell.value and "340+" in str(cell.value):
                cell.value = str(total)

    # Save
    output_path = "/Users/mrm/Projects/treasury-cd/docs/test-cases/Treasury-TestCases-BRD-v3.xlsx"
    wb.save(output_path)
    print(f"✅ Saved: {output_path}")
    print(f"📊 Total test cases: {total}")
    print(f"📋 Sheets: {len(sheets_config) + 1} (including guide)")


if __name__ == "__main__":
    main()